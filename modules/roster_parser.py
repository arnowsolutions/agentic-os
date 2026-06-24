#!/usr/bin/env python3
"""
Roster Parser — Parse monthly location-roster xlsx files into structured records.

Takes the raw Scheduling Grids (Nursing/Clerical sheets with date columns) and
outputs organized JSON per file with:
  - file_metadata (file name, date range, generated date)
  - sheets (Nursing, Clerical) each containing:
    - staff entries with {name, role, assignments: [{date, code, description}]}
    - date headers mapping column positions to dates

Usage:
    from modules.roster_parser import parse_roster
    result = parse_roster("path/to/schedule.xlsx", config)
"""

import json
import os
import re
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

CODE_LOOKUP = {
    # Common shift codes found in these grids
    "HT": "Working / Present",
    "X": "Off / Scheduled off", 
    "V": "Vacation",
    "VR": "Vacation Requested",
    "CE": "Conference",
    "CER": "Conference Requested",
    "DE": "Education Day",
    "H": "Pre-Holiday / Holiday",
    "HR": "Holiday Requested",
    "O": "Orientation",
    "Q": "Personal Day",
    "QR": "Personal Day Requested",
    "LOA": "Leave of Absence",
    "S": "Sick",
    "MAP": "MAP / Modified Assignment",
    "T2": "Training / T2",
    "T1": "Training / T1",
    "WF": "Work From Home / Remote",
    "KSP": "KSP / Special Assignment",
}

# ── Vacation parser (for Urology 2026 Approved Vacation.xlsm) ──

VAC_GROUP_LABELS = {
    "NPs, RNs & PAs": "NP/RN/PA",
    "LPNs, Secs & SW": "LPN/Secretary/SW",
    "RN - Nights": "RN (Nights)",
    "Clerical": "Clerical",
}

MONTH_NAMES = {
    "JANUARY":1,"FEBRUARY":2,"MARCH":3,"APRIL":4,"MAY":5,"JUNE":6,
    "JULY":7,"AUGUST":8,"SEPTEMBER":9,"OCTOBER":10,"NOVEMBER":11,"DECEMBER":12
}

def parse_roster(xlsx_path: str, config: Optional[dict] = None) -> dict:
    """
    Parse a monthly schedule xlsx into structured records.
    Returns {file, date_range, sheets: [{name, staff: [{name, role, assignments}]}]}
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    file_name = os.path.basename(xlsx_path)
    
    # Detect file type by name
    is_vacation = "vacation" in file_name.lower() or "vac" in file_name.lower()
    
    if is_vacation:
        sheets = _parse_vacation_file(wb, file_name)
        wb.close()
        return {
            "file": file_name,
            "date_range": {"start": "2026-01-01", "end": "2026-12-31", "label": "2026 Calendar Year"},
            "type": "vacation",
            "parsed_at": datetime.now().isoformat(),
            "sheets": sheets,
        }
    
    # Extract date range from filename e.g. "Urology 2026 Monthly Schedule (JUN07-JUL04)"
    date_range = _parse_date_range_from_filename(file_name)
    
    sheets = []
    for sheet_name in wb.sheetnames:
        if sheet_name.strip().upper() == "WINK":
            continue  # Skip legend sheet
        
        ws = wb[sheet_name]
        parsed = _parse_sheet(ws, sheet_name, file_name, date_range)
        if parsed is not None:
            sheets.append(parsed)
    
    wb.close()
    
    return {
        "file": file_name,
        "date_range": date_range,
        "parsed_at": datetime.now().isoformat(),
        "sheets": sheets,
    }


def _parse_date_range_from_filename(file_name: str) -> Dict:
    """
    Extract date range from filenames like:
    "Urology 2026 Monthly Schedule (JUN07-JUL04).xlsx"
    """
    match = re.search(r'\(([A-Z]{3})(\d{2})-([A-Z]{3})(\d{2})\)', file_name.upper())
    if match:
        start_month, start_day, end_month, end_day = match.groups()
        year_match = re.search(r'(20\d{2})', file_name)
        year = int(year_match.group(1)) if year_match else datetime.now().year
        
        months = {"JAN":1,"FEB":2,"MAR":3,"APR":4,"MAY":5,"JUN":6,
                  "JUL":7,"AUG":8,"SEP":9,"OCT":10,"NOV":11,"DEC":12}
        
        start_m = months.get(start_month, 1)
        end_m = months.get(end_month, 1)
        
        start_dt = f"{year}-{start_m:02d}-{int(start_day):02d}"
        end_dt = f"{year}-{end_m:02d}-{int(end_day):02d}"
        label = f"{start_month} {int(start_day)} - {end_month} {int(end_day)}, {year}"
        
        return {"start": start_dt, "end": end_dt, "label": label}
    
    return {"start": "", "end": "", "label": file_name}


def _parse_sheet(ws, sheet_name: str, file_name: str, date_range: dict) -> Optional[dict]:
    """Parse a single sheet (Nursing or Clerical) into staff entries."""
    import openpyxl
    
    # Step 1: Find the header row (has SUN./MON./TUE. etc)
    header_row_idx = None
    date_row_idx = None
    
    for r in range(1, min(ws.max_row + 1, 12)):
        val = str(ws.cell(r, 3).value or "").strip().upper()
        if val in ("SUN.", "SUN", "SUNDAY"):
            header_row_idx = r
            date_row_idx = r + 1
            break
    
    if header_row_idx is None:
        return None  # Can't find grid structure
    
    # Step 2: Build date-to-column mapping from header + date rows
    date_cols = []  # list of tuples: (col_idx, date_str)
    
    # The first 2 columns (A, B) are name/role
    for c in range(3, min(ws.max_column + 1, 32)):
        day_abbr = str(ws.cell(header_row_idx, c).value or "").strip().upper()[:3]
        date_val = ws.cell(date_row_idx, c).value
        
        if day_abbr and date_val:
            # Parse date string
            date_str = None
            if isinstance(date_val, datetime):
                date_str = date_val.strftime("%Y-%m-%d")
            elif isinstance(date_val, date):
                date_str = date_val.strftime("%Y-%m-%d")
            else:
                date_str = str(date_val).strip()[:10]
            
            if date_str:
                date_cols.append({"col": c, "day": day_abbr, "date": date_str})
    
    if not date_cols:
        return None
    
    # Step 3: Extract staff entries
    staff = []
    in_data_section = False
    total_keywords = ["TOTAL", "SUBTOTAL", "EXTENDER"]
    
    for r in range(header_row_idx + 2, min(ws.max_row + 1, 250)):
        name_val = str(ws.cell(r, 1).value or "").strip()
        role_val = str(ws.cell(r, 2).value or "").strip()
        
        # Skip empty rows, total rows, header rows
        upper_name = name_val.upper()
        if not name_val and not role_val:
            continue
        if any(kw in upper_name for kw in total_keywords):
            continue
        if upper_name in ("", "NURSING TEAM", "CLERICAL TEAM"):
            continue
        
        assignments = []
        for dc in date_cols:
            col = dc["col"]
            cell_val = str(ws.cell(r, col).value or "").strip()
            if cell_val:
                code = cell_val.split()[0]  # Take first word as code
                desc = CODE_LOOKUP.get(code, f"Unknown ({code})")
                # For present codes (HT), show it as working
                if code == "HT":
                    desc = "Working / Present"
                assignments.append({
                    "date": dc["date"],
                    "code": code,
                    "description": desc,
                })
        
        if name_val:
            entry = {
                "name": name_val,
                "role": role_val or None,
                "total_assignments": len(assignments),
                "working_days": sum(1 for a in assignments if a["code"] in ("HT", "") and a not in ("X", "V", "LOA", "S")),
                "assignments": assignments,
            }
            staff.append(entry)
    
    # Step 4: Also extract unassigned slots (row with name blank but role filled)
    # These are usually additional staff slots below totals
    
    return {
        "name": sheet_name,
        "staff_count": len(staff),
        "date_count": len(date_cols),
        "staff": staff,
    }


def query_location_roster(location: str = "", date_str: str = "") -> dict:
    """
    Query the parsed roster cache for who's at a location on a given date.
    
    location: optional filter (Nursing, Clerical, or empty for all)
    date_str: YYYY-MM-DD or empty for today
    
    Returns:
      {success, records: [{file, sheet, name, role, date, status, code}], note}
    """
    parsed_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 
                              "data", "location_rosters", "parsed")
    
    target_date = date_str or datetime.now().strftime("%Y-%m-%d")
    
    records = []
    for fname in sorted(os.listdir(parsed_dir)):
        if not fname.endswith(".json"):
            continue
        if "Contact" in fname:
            continue
        
        data = json.loads(open(os.path.join(parsed_dir, fname)).read())
        
        # Skip old-format files (pre-roster_parser — they have 'preview' key instead of 'sheets' list)
        if not isinstance(data.get("sheets"), list):
            continue
        
        # Skip if file doesn't cover the target date
        dr = data.get("date_range", {})
        dr_start = dr.get("start", "")
        dr_end = dr.get("end", "")
        if dr_start and dr_end:
            if target_date < dr_start or target_date > dr_end:
                continue
        
        for sheet in data.get("sheets", []):
            if isinstance(sheet, str):
                continue
            if location and location.lower() not in sheet.get("name", "").lower():
                continue
            
            for staff_member in sheet.get("staff", []):
                for assign in staff_member.get("assignments", []):
                    if assign.get("date") == target_date:
                        records.append({
                            "file": data.get("file", fname),
                            "sheet": sheet.get("name", ""),
                            "name": staff_member["name"],
                            "role": staff_member.get("role"),
                            "date": target_date,
                            "status": assign.get("description", ""),
                            "code": assign.get("code", ""),
                        })
    
    records.sort(key=lambda r: (r.get("sheet", ""), r.get("name", "")))
    
    location_label = location or "all departments"
    return {
        "success": True,
        "location": location_label,
        "date": target_date,
        "count": len(records),
        "records": records[:100],
        "note": f"Found {len(records)} staff assigned on {target_date} for {location_label}",
    }


def query_vacation(employee_name: str = "") -> dict:
    """
    Query approved vacation data for one person or all.
    
    Parsed vacation files have: sheets -> each with staff -> {name, group, weekly_assignments}
    
    Returns structured vacation data.
    """
    parsed_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)),
                              "data", "location_rosters", "parsed")
    
    records = []
    for fname in sorted(os.listdir(parsed_dir)):
        if not fname.endswith(".json"):
            continue
        data = json.loads(open(os.path.join(parsed_dir, fname)).read())
        if data.get("type") != "vacation":
            continue
        
        for sheet in data.get("sheets", []):
            if isinstance(sheet, str):
                continue
            group_label = sheet.get("group_label", sheet.get("name", ""))
            for staff_entry in sheet.get("staff", []):
                if isinstance(staff_entry, str):
                    continue
                n = staff_entry.get("name", "")
                if employee_name and employee_name.lower() not in n.lower():
                    continue
                records.append({
                    "name": n,
                    "group": group_label,
                    "hire_date": staff_entry.get("hire_date", ""),
                    "bargaining_date": staff_entry.get("bargaining_date", ""),
                    "vacation_days": staff_entry.get("vacation_days", 0),
                    "weeks": staff_entry.get("weeks", []),
                    "file": data.get("file", fname),
                })
    
    return {
        "success": True,
        "employee_filter": employee_name or "all",
        "count": len(records),
        "records": records,
        "note": f"Found {len(records)} employees" + (f" matching '{employee_name}'" if employee_name else " (all)"),
    }


def _parse_vacation_file(wb, file_name: str) -> list:
    """Parse vacation xlsm into sheets with staff vacation assignments."""
    sheets = []
    
    for sheet_name in wb.sheetnames:
        if sheet_name.strip().upper() == "WINK":
            continue
        
        ws = wb[sheet_name]
        group_label = VAC_GROUP_LABELS.get(sheet_name, sheet_name)
        
        # Find header row with months
        header_row = None
        for r in range(1, 6):
            v = str(ws.cell(r, 4).value or "").strip().upper()
            if v in MONTH_NAMES:
                header_row = r
                break
        
        if header_row is None:
            continue
        
        # Build month->column mapping
        # Format: Row has alternating JANUARY, FEBRUARY, etc. with 4 columns per month
        month_cols = []  # [(month_number, col_start), ...]
        for c in range(4, min(ws.max_column + 1, 62)):
            v = str(ws.cell(header_row, c).value or "").strip().upper()
            if v in MONTH_NAMES:
                month_num = MONTH_NAMES[v]
                # Each month spans 4 columns (week dates)
                year = 2026
                month_cols.append({
                    "month": month_num,
                    "col_start": c,
                    "year": year,
                })
        
        # Parse staff rows
        staff_list = []
        for r in range(header_row + 2, min(ws.max_row + 1, 300)):
            name_val = str(ws.cell(r, 1).value or "").strip()
            hire_val = ws.cell(r, 2).value
            
            if not name_val:
                continue
            if any(kw in name_val.upper() for kw in ["TOTAL", "SUBTOTAL", "ASSOCIATE"]):
                continue
            
            # Parse hire/bargaining date
            hire_date = ""
            if isinstance(hire_val, datetime):
                hire_date = hire_val.strftime("%Y-%m-%d")
            elif hire_val:
                hire_date = str(hire_val)[:10]
            
            weeks = []
            for mc in month_cols:
                month_num = mc["month"]
                col_start = mc["col_start"]
                year = mc["year"]
                # 4 week columns per month
                for wi in range(4):
                    c = col_start + wi
                    if c > ws.max_column:
                        break
                    cell_v = str(ws.cell(r, c).value or "").strip()
                    if cell_v:
                        weeks.append({"month": month_num, "week": wi + 1, "code": cell_v})
            
            staff_list.append({
                "name": name_val,
                "hire_date": hire_date,
                "group": group_label,
                "vacation_days": len([w for w in weeks if w["code"].upper() == "V"]),
                "weeks": weeks,
            })
        
        sheets.append({
            "name": sheet_name,
            "group_label": group_label,
            "staff_count": len(staff_list),
            "staff": staff_list,
        })
    
    return sheets


def rebuild_parsed_cache(force: bool = False) -> dict:
    """Reparse all raw roster files into the parsed cache."""
    raw_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)),
                           "data", "location_rosters", "raw")
    parsed_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)),
                              "data", "location_rosters", "parsed")
    
    os.makedirs(parsed_dir, exist_ok=True)
    
    synced = 0
    errors = []
    
    for fname in sorted(os.listdir(raw_dir)):
        if not fname.endswith(".xlsx") or "Contact" in fname:
            continue
        
        raw_path = os.path.join(raw_dir, fname)
        parsed_path = os.path.join(parsed_dir, f"{fname}.json")
        
        try:
            result = parse_roster(raw_path)
            with open(parsed_path, "w") as f:
                json.dump(result, f, indent=2, default=str)
            synced += 1
        except Exception as e:
            errors.append(f"{fname}: {str(e)[:200]}")
    
    return {"success": True, "synced": synced, "errors": errors}


def staff_at_location(location: str, date_str: str = "") -> dict:
    """
    Query who's working/scheduled at a physical office location on a given date.

    This wraps query_location_roster with data_status semantics so Vapi
    can phrase responses naturally.

    Returns:
      {success, data_status, records, location, date, period_end, known_locations}
    
    data_status values:
      "ok"                    — records returned
      "no_assignments"        — location found but no one assigned on that date
      "no_data_for_date"      — date outside any loaded roster's range
      "unknown_location"      — location not recognized
      "rosters_not_synced"    — parsed directory empty / not yet synced
    """
    parsed_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)),
                              "data", "location_rosters", "parsed")

    target_date = date_str or datetime.now().strftime("%Y-%m-%d")

    if not os.path.isdir(parsed_dir) or not os.listdir(parsed_dir):
        return {
            "success": True,
            "data_status": "rosters_not_synced",
            "records": [],
            "location": location,
            "date": target_date,
            "known_locations": [],
        }

    # Collect known locations (unique sheet names from all parsed files)
    known_locations = set()
    for fname in sorted(os.listdir(parsed_dir)):
        if not fname.endswith(".json") or "Contact" in fname:
            continue
        data = json.loads(open(os.path.join(parsed_dir, fname)).read())
        if not isinstance(data.get("sheets"), list):
            continue
        for sheet in data.get("sheets", []):
            if isinstance(sheet, str):
                continue
            known_locations.add(sheet.get("name", "").strip())

    known_locations = sorted(known_locations)

    # Check if location is recognized
    loc_lower = location.lower().strip()
    matching_sites = [s for s in known_locations if loc_lower in s.lower()]
    if not matching_sites:
        return {
            "success": True,
            "data_status": "unknown_location",
            "records": [],
            "location": location,
            "date": target_date,
            "known_locations": known_locations,
        }

    # Check if date falls within any loaded roster
    date_in_any_roster = False
    newest_period_end = ""
    for fname in sorted(os.listdir(parsed_dir)):
        if not fname.endswith(".json") or "Contact" in fname:
            continue
        data = json.loads(open(os.path.join(parsed_dir, fname)).read())
        dr = data.get("date_range", {})
        dr_start = dr.get("start", "")
        dr_end = dr.get("end", "")
        if dr_start and dr_end:
            if dr_end > newest_period_end:
                newest_period_end = dr_end
            if dr_start <= target_date <= dr_end:
                date_in_any_roster = True

    if not date_in_any_roster:
        return {
            "success": True,
            "data_status": "no_data_for_date",
            "records": [],
            "location": location,
            "date": target_date,
            "period_end": newest_period_end,
            "known_locations": known_locations,
        }

    # Query the roster for this location + date
    raw = query_location_roster(location, target_date)

    if raw.get("count", 0) == 0:
        return {
            "success": True,
            "data_status": "no_assignments",
            "records": [],
            "location": location,
            "date": target_date,
            "known_locations": known_locations,
        }

    # Build clean people list
    people = []
    for r in raw.get("records", []):
        people.append({
            "name": r.get("name", ""),
            "role": r.get("role", ""),
            "status": r.get("status", ""),
        })

    working_statuses = {"working / present", "work from home / remote", "orientation", "training"}
    working_people = [p for p in people if p.get("status", "").lower().strip() in working_statuses]

    if len(working_people) == 0:
        return {
            "success": True,
            "data_status": "no_assignments",
            "records": [],
            "location": location,
            "date": target_date,
            "known_locations": known_locations,
        }

    return {
        "success": True,
        "data_status": "ok",
        "records": working_people,
        "count": len(working_people),
        "location": location,
        "date": target_date,
        "known_locations": known_locations,
    }


# ── CLI entry point for testing ──────────────────────────────────────
if __name__ == "__main__":
    import sys
    action = sys.argv[1] if len(sys.argv) > 1 else "rebuild"
    
    if action == "rebuild":
        result = rebuild_parsed_cache()
        print(json.dumps(result, indent=2))
    elif action == "query":
        loc = sys.argv[2] if len(sys.argv) > 2 else ""
        dt = sys.argv[3] if len(sys.argv) > 3 else ""
        result = query_location_roster(loc, dt)
        print(json.dumps(result, indent=2, default=str)[:2000])
    elif action == "parse":
        path = sys.argv[2]
        result = parse_roster(path)
        print(json.dumps(result, indent=2, default=str)[:3000])
