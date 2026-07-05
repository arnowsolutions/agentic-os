#!/usr/bin/env python3
"""
Drive Sync — Pull location-roster files from Google Drive into a local cache.

Single-purpose module that syncs Excel/PDF rosters from Drive into
data/location_rosters/ so the assistant can answer "who's at [location]
on [date]" without reaching Drive at query time.

Usage:
    from modules import drive_sync
    result = drive_sync.sync_location_rosters()
    result = drive_sync.sync_location_rosters(force=True)
"""

import json
import os
import re
import sys
import traceback
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

# ── Path resolution (works on VPS and container) ──────────────────────
_THIS_DIR = Path(__file__).resolve().parent
_WORKSPACE = os.environ.get("WORKSPACE") or str(_THIS_DIR.parent)
_CONFIG_PATH = os.path.join(_WORKSPACE, "data", "drive_sync_config.json")
_CACHE_ROOT = os.path.join(os.path.expanduser("~/.hermes/data"), "location_rosters")

# ── Initialize GoogleWorkspace (lazy) ─────────────────────────────────
_gw = None


def _get_workspace():
    """Lazy-init GoogleWorkspace. Returns None if not configured."""
    global _gw
    if _gw is not None:
        return _gw
    try:
        sys.path.insert(0, _WORKSPACE)
        from modules.google_workspace import GoogleWorkspace
        # Use token from shared workspace volume (accessible on both container and VPS)
        tok = os.path.join(_WORKSPACE, "data", "google_token.json")
        cred = os.path.join(_WORKSPACE, "data", "google_credentials.json")
        if not os.path.exists(tok):
            # Fallback to home dir
            tok = os.path.expanduser("~/.hermes/google_token.json")
            cred = os.path.expanduser("~/.hermes/google_credentials.json")
        _gw = GoogleWorkspace(
            user_id="urologyresidency",
            token_path=tok,
            credentials_path=cred,
        )
        # Test credentials
        _gw._get_credentials()
        return _gw
    except RuntimeError as e:
        _gw = None
        return None
    except Exception as e:
        _gw = None
        return None


# ── Public API ────────────────────────────────────────────────────────


def sync_location_rosters(force: bool = False) -> dict:
    """
    Sync location-roster files from Drive to local cache.

    Returns:
        {success, synced, skipped, errors, files, reason (if failed)}
    """
    config = _load_config()
    if not config:
        return {"success": False, "reason": "config_missing",
                "action": f"Create {_CONFIG_PATH} from .example"}

    folder_id = config.get("rosters_folder_id", "")
    search_query = config.get("rosters_search_query", "")
    if not folder_id and not search_query:
        return {"success": False, "reason": "no_config",
                "action": "Set rosters_folder_id or rosters_search_query in drive_sync_config.json"}

    gw = _get_workspace()
    if gw is None:
        return {"success": False, "reason": "google_oauth_not_configured",
                "action": "Run python3 /workspace/auth_google.py"}

    cache_dir = _CACHE_ROOT
    raw_dir = os.path.join(cache_dir, "raw")
    parsed_dir = os.path.join(cache_dir, "parsed")
    manifest_path = os.path.join(cache_dir, "_manifest.json")

    os.makedirs(raw_dir, exist_ok=True)
    os.makedirs(parsed_dir, exist_ok=True)

    # Load existing manifest
    manifest = _load_manifest(manifest_path)

    # List Drive files (by folder or search query)
    # List Drive files (by folder or search query)
    extra_files = config.get("extra_files", [])
    
    try:
        if folder_id and not folder_id.startswith("YOUR_"):
            drive_result = gw.list_drive_files(
                user_id="urologyresidency",
                folder_id=folder_id,
            )
        elif search_query:
            raw = gw.drive.files().list(
                q=search_query,
                fields="files(id, name, mimeType, modifiedTime)",
                supportsAllDrives=True,
                includeItemsFromAllDrives=True,
                pageSize=50
            ).execute()
            drive_result = {"successful": True, "data": raw}
        else:
            drive_result = {"successful": True, "data": {"files": []}}
    except Exception as e:
        return {"success": False, "reason": "drive_list_failed",
                "error": str(e)[:500]}

    if not drive_result.get("successful"):
        return {"success": False, "reason": "drive_list_failed",
                "error": drive_result.get("error", "unknown")[:500]}

    drive_files = drive_result.get("data", {}).get("files", [])
    if not drive_files:
        return {"success": True, "synced": 0, "skipped": 0, "errors": [],
                "files": [], "note": "No files found in Drive folder"}

    synced = 0
    skipped = 0
    errors: List[str] = []
    file_results: List[dict] = []

    for df in drive_files:
        file_id = df.get("id", "")
        file_name = df.get("name", "")
        if not file_id or not file_name:
            continue

        # Only process xlsx and pdf files
        if not (file_name.lower().endswith(".xlsx") or
                file_name.lower().endswith(".pdf")):
            continue

        # Get modifiedTime (not included in list_drive_files fields)
        drive_modified = _get_file_modified(gw, file_id)
        if drive_modified is None:
            errors.append(f"{file_name}: Could not get modifiedTime")
            continue

        # Check manifest for existing entry
        existing = manifest.get(file_id, {})
        cached_mtime = existing.get("drive_modified_time", "")

        if not force and cached_mtime == drive_modified:
            skipped += 1
            file_results.append({
                "file_id": file_id, "name": file_name,
                "status": "unchanged",
                "drive_modified_time": drive_modified,
            })
            continue

        # Download file
        raw_path = os.path.join(raw_dir, file_name)
        try:
            dl_result = gw.download_drive_file(
                user_id="urologyresidency",
                file_id=file_id,
                dest_path=raw_path,
            )
            if not dl_result.get("successful"):
                errors.append(f"{file_name}: Download failed - "
                              f"{dl_result.get('error', 'unknown')[:200]}")
                file_results.append({
                    "file_id": file_id, "name": file_name,
                    "status": "error",
                    "drive_modified_time": drive_modified,
                })
                continue
        except Exception as e:
            errors.append(f"{file_name}: Download exception - {str(e)[:200]}")
            continue

        # Parse (delegated to roster_parser — placeholder for now)
        parsed_path = os.path.join(parsed_dir, f"{file_name}.json")
        parse_status = "skipped"
        parse_error = None
        try:
            # Try parsing via roster_parser if available
            parsed = _try_parse(raw_path, file_name, config)
            if parsed is not None:
                os.makedirs(os.path.dirname(parsed_path), exist_ok=True)
                with open(parsed_path, "w") as f:
                    json.dump(parsed, f, indent=2, default=str)
                parse_status = "ok"
        except Exception as e:
            parse_status = "error"
            parse_error = str(e)[:300]

        # Update manifest
        manifest[file_id] = {
            "file_id": file_id,
            "name": file_name,
            "drive_modified_time": drive_modified,
            "local_path": raw_path,
            "parsed_path": parsed_path if parse_status == "ok" else "",
            "parse_status": parse_status,
            "error": parse_error,
            "last_synced": datetime.now(timezone.utc).isoformat(),
        }

        status = "synced" if parse_status == "ok" else parse_status
        file_results.append({
            "file_id": file_id, "name": file_name,
            "status": status,
            "drive_modified_time": drive_modified,
        })
        synced += 1

    # Process extra files from config
    for extra in extra_files:
        file_id = extra.get("drive_file_id", "")
        file_name = extra.get("name", "") + "." + (extra.get("type", "xlsx")) if extra.get("name") else ""
        if not file_id:
            continue
        
        # Derive filename from the file_id
        try:
            fm = gw.drive.files().get(fileId=file_id, fields="name", supportsAllDrives=True).execute()
            file_name = fm.get("name", file_name)
        except Exception:
            pass
        
        # Check manifest
        existing = manifest.get(file_id, {})
        cached_mtime = existing.get("drive_modified_time", "")
        
        # Get modified time
        drive_modified = _get_file_modified(gw, file_id)
        
        if not force and cached_mtime == drive_modified:
            skipped += 1
            file_results.append({"file_id": file_id, "name": file_name, "status": "unchanged"})
            continue
        
        # Download
        raw_path = os.path.join(raw_dir, file_name)
        try:
            dl = gw.download_drive_file(user_id="urologyresidency", file_id=file_id, dest_path=raw_path)
            if not dl.get("successful"):
                errors.append(f"{file_name}: extra DL failed - {dl.get('error','')[:200]}")
                continue
        except Exception as e:
            errors.append(f"{file_name}: extra DL exception - {str(e)[:200]}")
            continue
        
        # Parse
        parsed_path = os.path.join(parsed_dir, f"{file_name}.json")
        parse_status = "skipped"
        parse_error = None
        try:
            parsed = _try_parse(raw_path, file_name, config)
            if parsed:
                os.makedirs(os.path.dirname(parsed_path), exist_ok=True)
                with open(parsed_path, "w") as f:
                    json.dump(parsed, f, indent=2, default=str)
                parse_status = "ok"
        except Exception as e:
            parse_status = "error"
            parse_error = str(e)[:300]
        
        manifest[file_id] = {
            "file_id": file_id, "name": file_name,
            "drive_modified_time": drive_modified,
            "local_path": raw_path,
            "parsed_path": parsed_path if parse_status == "ok" else "",
            "parse_status": parse_status, "error": parse_error,
            "last_synced": datetime.now(timezone.utc).isoformat(),
        }
        synced += 1
        file_results.append({"file_id": file_id, "name": file_name, "status": parse_status})
    
    # Save manifest
    _save_manifest(manifest_path, manifest)

    return {
        "success": True,
        "synced": synced,
        "skipped": skipped,
        "errors": errors,
        "files": file_results,
    }


def get_sync_status() -> dict:
    """Read manifest and report per-file freshness."""
    manifest_path = os.path.join(_CACHE_ROOT, "_manifest.json")
    manifest = _load_manifest(manifest_path)

    files_status = []
    now = datetime.now(timezone.utc)
    for fid, entry in manifest.items():
        last_synced = entry.get("last_synced", "")
        age_days = None
        if last_synced:
            try:
                dt = datetime.fromisoformat(last_synced)
                age_days = round((now - dt).total_seconds() / 86400, 1)
            except (ValueError, TypeError):
                pass
        files_status.append({
            "name": entry.get("name", fid),
            "file_id": fid,
            "last_synced": last_synced,
            "age_days": age_days,
            "parse_status": entry.get("parse_status", "unknown"),
            "drive_modified_time": entry.get("drive_modified_time", ""),
        })

    files_status.sort(key=lambda f: f.get("name", ""))

    return {
        "manifest_file": manifest_path,
        "total_files": len(files_status),
        "files": files_status,
    }


# ── Internal helpers ──────────────────────────────────────────────────


def _load_config() -> Optional[dict]:
    if not os.path.exists(_CONFIG_PATH):
        return None
    try:
        return json.loads(open(_CONFIG_PATH).read())
    except (json.JSONDecodeError, OSError):
        return None


def _load_manifest(path: str) -> dict:
    if os.path.exists(path):
        try:
            return json.loads(open(path).read())
        except (json.JSONDecodeError, OSError):
            pass
    return {}


def _save_manifest(path: str, data: dict):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w") as f:
        json.dump(data, f, indent=2, default=str)


def _get_file_modified(gw, file_id: str) -> Optional[str]:
    """Get modifiedTime for a Drive file (not in list_drive_files fields)."""
    try:
        result = (
            gw.drive.files()
            .get(fileId=file_id, fields="modifiedTime",
                 supportsAllDrives=True)
            .execute()
        )
        return result.get("modifiedTime")
    except Exception:
        return None


def _try_parse(raw_path: str, file_name: str, config: dict) -> Optional[Any]:
    """
    Attempt to parse a roster file using roster_parser.
    """
    try:
        from modules.roster_parser import parse_roster
        result = parse_roster(raw_path, config)
        return result
    except Exception as e:
        # Fallback: basic structure extraction
        if file_name.lower().endswith(".xlsx"):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(raw_path, data_only=True)
                sheets = {}
                for sn in wb.sheetnames:
                    ws = wb[sn]
                    rows = []
                    for row in ws.iter_rows(min_row=1,
                                             max_row=min(ws.max_row, 200),
                                             values_only=True):
                        rows.append([str(v) if v is not None else "" for v in row])
                    sheets[sn] = rows[:50]
                wb.close()
                return {"file": file_name, "sheets": list(wb.sheetnames),
                        "preview": sheets, "parser_version": "fallback"}
            except Exception:
                pass
        return None
