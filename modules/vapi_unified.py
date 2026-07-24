"""Unified data layer for Vapi voice queries.
   Reads from local xlsx + CSV files for schedule, GME, QGenda daily assignments.
   No Drive/Composio dependency needed.
"""
import csv
import json
import os
import sqlite3
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

# ── Paths — resolve relative to workspace root ──────────
_THIS_DIR = Path(__file__).resolve().parent  # modules/
_WORKSPACE = os.environ.get("WORKSPACE") or str(_THIS_DIR.parent.parent)  # agentic-os/../.. = /workspace
SCHEDULE_PATH = os.path.join(_WORKSPACE, "Call_Schedule_Q3_Q4_2026.xlsx")
GME_PATH = os.path.join(_WORKSPACE, "Resident_Trackers2025-2026.xlsx")
QGENDA_PATH = os.path.join(_WORKSPACE, "repos/qgenda/data/Montefiore_Medical_Center_-_Urology_Schedule_Export_1-1-2026_to_12-31-2026.csv")
REIMBURSEMENT_DB = os.path.join(_WORKSPACE, "repos/reimbursement/reimbursement.db")
STAFF_PATH = os.path.join(_WORKSPACE, "repos/sick-call-line/data/associates.csv")
GME_CAP = 1250.0

# ── Helpers ────────────────────────────────────────────
_qgenda_cache = None
_schedule_cache = None
_staff_cache = None


# ============================================================================
# QGENDA — Full-year daily assignments (4,272 rows)
# "Where is Dr. Sankin today?"  "What clinic am I at tomorrow?"
# ============================================================================

def _load_qgenda():
    global _qgenda_cache
    if _qgenda_cache is not None:
        return _qgenda_cache
    
    if not os.path.exists(QGENDA_PATH):
        return {"error": "QGenda data not available"}
    
    with open(QGENDA_PATH, encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        rows = list(reader)
    
    # Index by person (multiple name forms) and by date
    by_person = defaultdict(list)
    by_date = defaultdict(list)
    by_task = defaultdict(list)
    all_names = set()
    
    for row in rows:
        first = row.get("Staff First Name", "").strip()
        last = row.get("Staff Last Name", "").strip()
        email = row.get("Staff Email", "").strip()
        date_str = row.get("Schedule Date", "").strip()
        task = row.get("Task Name", "").strip()
        
        full_name = f"{first} {last}"
        key = full_name.lower()
        
        entry = {"name": full_name, "first": first, "last": last, "email": email, "date": date_str, "task": task}
        
        # Index under multiple keys for fuzzy lookup
        by_person[key].append(entry)
        by_person[first.lower()].append(entry)
        by_person[last.lower()].append(entry)
        all_names.add(full_name)
        
        by_date[date_str].append(entry)
        by_task[task].append(entry)
    
    _qgenda_cache = {
        "by_person": dict(by_person),
        "by_date": dict(by_date),
        "by_task": dict(by_task),
        "all_names": sorted(all_names),
        "total_rows": len(rows),
    }
    return _qgenda_cache


def qgenda_today(name=None):
    """Get today's schedule. If name provided, filter for that person."""
    data = _load_qgenda()
    if isinstance(data, dict) and "error" in data:
        return data
    
    today = date.today()
    qfmt = f"{today.month:02d}-{today.day:02d}-{str(today.year)[-2:]}"
    
    if name:
        return qgenda_person_day(name, qfmt)
    
    entries = data["by_date"].get(qfmt, [])
    if not entries:
        return {"date": qfmt, "note": "No schedule data for today (weekend or holiday)", "entries": []}
    
    # Group by task
    by_task = defaultdict(list)
    for e in entries:
        by_task[e["task"]].append(e["name"])
    
    return {
        "date": qfmt,
        "total": len(entries),
        "summary": {task: {"count": len(names), "people": names[:10]} for task, names in sorted(by_task.items(), key=lambda x: -len(x[1]))},
        "entries": entries[:30],
    }


def qgenda_person_day(name: str, date_str: str = None):
    """Get a person's assignments on a specific date."""
    data = _load_qgenda()
    if isinstance(data, dict) and "error" in data:
        return data
    
    if date_str is None:
        today = date.today()
        date_str = f"{today.month:02d}-{today.day:02d}-{str(today.year)[-2:]}"
    
    name_lower = name.lower().strip()
    results = []
    
    # Check all indices for this name
    for key in [name_lower, name_lower.replace("dr. ", ""), name_lower.replace("dr ", "")]:
        entries = data["by_person"].get(key, [])
        for e in entries:
            if e["date"] == date_str:
                results.append(e)
    
    return {"name": name, "date": date_str, "assignments": results, "count": len(results)}


def qgenda_today_task(task_name: str):
    """Get all people doing a specific task today. 'Who's at the Stone Clinic?'"""
    data = _load_qgenda()
    if isinstance(data, dict) and "error" in data:
        return data
    
    today = date.today()
    qfmt = f"{today.month}-{today.day}-{str(today.year)[-2:]}"
    
    task_lower = task_name.lower().strip()
    matches = []
    for e in data["by_date"].get(qfmt, []):
        if task_lower in e["task"].lower():
            matches.append(e)
    
    return {"task": task_name, "date": qfmt, "people": [m["name"] for m in matches], "count": len(matches)}


def qgenda_person_upcoming(name: str, days: int = 7):
    """Get a person's upcoming schedule for the next N days."""
    data = _load_qgenda()
    if isinstance(data, dict) and "error" in data:
        return data
    
    name_lower = name.lower().strip()
    today = date.today()
    end = today + timedelta(days=days)
    
    results = []
    for key in [name_lower, name_lower.replace("dr. ", ""), name_lower.replace("dr ", "")]:
        for e in data["by_person"].get(key, []):
            # Parse date
            parts = e["date"].split("-")
            if len(parts) == 3:
                try:
                    d = date(2000 + int(parts[2]), int(parts[0]), int(parts[1]))
                    if today <= d <= end:
                        results.append({"date": e["date"], "day": d.strftime("%A"), "task": e["task"]})
                except:
                    pass
    
    results.sort(key=lambda x: x["date"])
    return {"name": name, "days": days, "assignments": results, "count": len(results)}


# ============================================================================
# CALL SCHEDULE — Attending call coverage (from xlsx)
# ============================================================================
# NOTE: call_today() and call_person() have been removed.
# Use vapi_data.schedule_today() and vapi_data.schedule_person() instead.
# schedule_by_date(), call_weekend(), and call_month() remain here.

def _load_schedule():
    global _schedule_cache
    if _schedule_cache is not None:
        return _schedule_cache
    
    import openpyxl
    wb = openpyxl.load_workbook(SCHEDULE_PATH, data_only=True)
    data = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for r in range(2, ws.max_row + 1):
            dv = ws.cell(r, 1).value
            if isinstance(dv, datetime):
                ds = dv.strftime("%Y-%m-%d")
            elif isinstance(dv, date):
                ds = dv.strftime("%Y-%m-%d")
            else:
                ds = str(dv or "").strip()[:10]
            rows.append({
                "date": ds,
                "day": str(ws.cell(r, 2).value or ""),
                "primary": str(ws.cell(r, 3).value or "").strip(),
                "backup": str(ws.cell(r, 4).value or "").strip(),
                "peds": str(ws.cell(r, 5).value or "").strip(),
                "primary_clean": str(ws.cell(r, 3).value or "").strip().split("(")[0].strip().split(",")[0].strip(),
                "backup_clean": str(ws.cell(r, 4).value or "").strip().split("(")[0].strip().split(",")[0].strip(),
                "peds_clean": str(ws.cell(r, 5).value or "").strip().split("(")[0].strip().split(",")[0].strip(),
            })
        data[sheet_name.lower()] = rows
    _schedule_cache = data
    return data


def call_weekend():
    """Get the weekend call block (Fri-Sun; Monday only when it's a holiday)."""
    sched = _load_schedule()
    dates = _weekend_dates()
    result = {"dates": dates}
    for campus, rows in sched.items():
        campus_data = []
        for row in rows:
            if row["date"] in dates:
                campus_data.append({
                    "date": row["date"], "day": row["day"],
                    "primary": row["primary"] or "—",
                    "backup": row["backup"] or "—",
                    "peds": row["peds"] or "—",
                })
        result[campus.lower()] = campus_data
    return result


def call_month(name: str):
    """Get a person's upcoming 60-day call schedule."""
    sched = _load_schedule()
    name = name.lower().strip()
    today_str = date.today().strftime("%Y-%m-%d")
    cutoff = (date.today() + timedelta(days=60)).strftime("%Y-%m-%d")
    results = []
    for campus, rows in sched.items():
        for row in rows:
            if row["date"] < today_str or row["date"] > cutoff:
                continue
            for role in ["primary_clean", "backup_clean", "peds_clean"]:
                val = row[role]
                if val and name in val.lower():
                    results.append({
                        "campus": campus.title(), "date": row["date"],
                        "day": row["day"], "role": role.replace("_clean", ""),
                    })
                    break
    return {"results": results, "total": len(results)}


import re as _re

_MONTH_MAP = {
    "jan": 1, "january": 1, "feb": 2, "february": 2, "mar": 3, "march": 3,
    "apr": 4, "april": 4, "may": 5, "jun": 6, "june": 6, "jul": 7, "july": 7,
    "aug": 8, "august": 8, "sep": 9, "sept": 9, "september": 9, "oct": 10, "october": 10,
    "nov": 11, "november": 11, "dec": 12, "december": 12,
}


def _normalize_date(date_str: str) -> str | None:
    """Convert any date string the LLM might pass into YYYY-MM-DD.

    Handles: 'today', 'tomorrow', 'yesterday', 'next Monday',
    'July 2', 'Jul 2', '7/2/2026', '2026-7-2', '07-02-26', etc.
    Returns None if it cannot parse the input.
    """
    if not date_str or not date_str.strip():
        return None
    raw = date_str.strip()
    low = raw.lower()

    # Relative keywords
    today = date.today()
    if low in ("today", "todays", "tonight"):
        return today.strftime("%Y-%m-%d")
    if low in ("tomorrow", "tomorrows", "tmrw"):
        return (today + timedelta(days=1)).strftime("%Y-%m-%d")
    if low in ("yesterday",):
        return (today - timedelta(days=1)).strftime("%Y-%m-%d")
    if low in ("day after tomorrow",):
        return (today + timedelta(days=2)).strftime("%Y-%m-%d")

    # "next <weekday>" — upcoming occurrence of that weekday
    weekday_map = {"monday": 0, "tuesday": 1, "wednesday": 2, "thursday": 3,
                   "friday": 4, "saturday": 5, "sunday": 6}
    for wname, wnum in weekday_map.items():
        if low == f"next {wname}" or low == wname:
            days_ahead = (wnum - today.weekday()) % 7
            if days_ahead == 0:
                days_ahead = 7
            return (today + timedelta(days=days_ahead)).strftime("%Y-%m-%d")
        if low == f"this {wname}":
            days_ahead = (wnum - today.weekday()) % 7
            return (today + timedelta(days=days_ahead)).strftime("%Y-%m-%d")

    # Already correct format: YYYY-MM-DD (padded or not)
    m = _re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})$", raw)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return date(y, mo, d).strftime("%Y-%m-%d")
        except ValueError:
            return None

    # M/D/YYYY or M-D-YYYY or MM/DD/YYYY
    m = _re.match(r"^(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})$", raw)
    if m:
        mo, d, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return date(y, mo, d).strftime("%Y-%m-%d")
        except ValueError:
            return None

    # MM-DD-YY (2-digit year, QGenda style)
    m = _re.match(r"^(\d{1,2})-(\d{1,2})-(\d{2})$", raw)
    if m:
        mo, d, yr = int(m.group(1)), int(m.group(2)), int(m.group(3))
        y = 2000 + yr if yr < 50 else 1900 + yr
        try:
            return date(y, mo, d).strftime("%Y-%m-%d")
        except ValueError:
            return None

    # "July 2", "Jul 2", "July 2 2026", "Jul 2 2026"
    m = _re.match(r"^([A-Za-z]+)\s+(\d{1,2})(?:\s*,?\s*(\d{4}))?", raw)
    if m:
        month_name = m.group(1).lower()
        if month_name in _MONTH_MAP:
            mo = _MONTH_MAP[month_name]
            d = int(m.group(2))
            y = int(m.group(3)) if m.group(3) else today.year
            try:
                return date(y, mo, d).strftime("%Y-%m-%d")
            except ValueError:
                return None

    # "2 July", "2 Jul 2026"
    m = _re.match(r"^(\d{1,2})\s+([A-Za-z]+)(?:\s*,?\s*(\d{4}))?", raw)
    if m:
        d = int(m.group(1))
        month_name = m.group(2).lower()
        if month_name in _MONTH_MAP:
            mo = _MONTH_MAP[month_name]
            y = int(m.group(3)) if m.group(3) else today.year
            try:
                return date(y, mo, d).strftime("%Y-%m-%d")
            except ValueError:
                return None

    # Could not parse
    return None


def schedule_by_date(date_str: str) -> dict:
    """Get call coverage for a specific date across all campuses.

    Accepts any date format the LLM might pass: 'today', 'tomorrow',
    'July 2', '2026-7-2', '7/2/2026', '07-02-26', etc.
    Normalizes to YYYY-MM-DD before matching against the Excel data.
    """
    sched = _load_schedule()
    normalized = _normalize_date(date_str)
    if normalized is None:
        return {
            "date": date_str,
            "note": (
                f"I couldn't understand the date '{date_str}'. "
                "Please say it as a month and day, like 'July 2' or 'July 2 2026', "
                "or use YYYY-MM-DD format."
            ),
        }
    result = {"date": normalized, "campuses": {}}
    for campus, rows in sched.items():
        for row in rows:
            if row["date"] == normalized:
                result["campuses"][campus.title()] = {
                    "primary": row["primary"] or "—",
                    "backup": row["backup"] or "—",
                    "peds": row["peds"] or "—",
                }
                break
    return result if result["campuses"] else {
        "date": normalized,
        "note": "No data for this date. Schedule covers July 2026 through January 2027.",
    }


# ============================================================================
# GME — Resident reimbursement balances
# ============================================================================

def gme_balance(name: str):
    """Get a resident's GME balance. Case-insensitive, partial match."""
    if not os.path.exists(GME_PATH):
        return {"error": "GME data file not found"}
    
    import openpyxl
    try:
        wb = openpyxl.load_workbook(GME_PATH, data_only=True)
    except Exception as e:
        return {"error": f"Could not open GME file: {e}"}
    
    ws = wb["Sheet 2025-2026"]
    name_lower = name.lower().strip()
    
    total = 0.0
    txn_count = 0
    transactions = []
    matched_name = None
    
    for r in range(2, ws.max_row + 1):
        row_name = str(ws.cell(r, 1).value or "").strip()
        if not row_name:
            continue
        rl = row_name.lower()
        
        if name_lower == rl or name_lower in rl or rl in name_lower:
            if matched_name is None:
                matched_name = row_name
            amount = ws.cell(r, 5).value
            desc = str(ws.cell(r, 4).value or "")[:60]
            ds = ws.cell(r, 3).value or ws.cell(r, 2).value
            date_str = ds.strftime("%Y-%m-%d") if isinstance(ds, datetime) else str(ds or "?")
            
            if amount and isinstance(amount, (int, float)):
                total += abs(amount)
                txn_count += 1
                transactions.append({"date": date_str, "description": desc, "amount": round(abs(amount), 2)})
    
    if matched_name is None:
        return {"error": f"No GME records found for '{name}'"}
    
    remaining = max(0, GME_CAP - min(total, GME_CAP))
    return {
        "name": matched_name, "total_spent": round(total, 2),
        "remaining": round(remaining, 2), "cap": GME_CAP,
        "transaction_count": txn_count,
        "recent_transactions": sorted(transactions, key=lambda x: x["date"], reverse=True)[:5],
    }


# ============================================================================
# STAFF DIRECTORY (from sick-call associates)
# ============================================================================

def _load_staff():
    global _staff_cache
    if _staff_cache is not None:
        return _staff_cache
    
    if not os.path.exists(STAFF_PATH):
        return {"error": "Staff data not found"}
    
    with open(STAFF_PATH, encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        staff = []
        for row in reader:
            name = row.get("Employee Name", "").strip()
            email = row.get("Email", "").strip()
            phone = row.get("Phone", "").strip()
            location = row.get("Location Code", "").strip()
            emp_id = row.get("Employee ID", "").strip()
            
            # Parse "Last, First" format
            parts = name.split(",")
            if len(parts) >= 2:
                first = parts[1].strip().split()[0] if parts[1].strip() else ""
                last = parts[0].strip()
                display = f"{first} {last}"
            else:
                first = ""
                last = name
                display = name
            
            staff.append({
                "display_name": display,
                "first": first.lower(),
                "last": last.lower(),
                "full_lower": display.lower(),
                "email": email,
                "phone": phone,
                "location": location,
                "employee_id": emp_id,
            })
    
    _staff_cache = staff
    return staff


def staff_lookup(name: str):
    """Find a staff member by name."""
    staff = _load_staff()
    if isinstance(staff, dict) and "error" in staff:
        return staff
    
    name_lower = name.lower().strip()
    results = []
    for s in staff:
        if name_lower in s["full_lower"] or name_lower in s["first"] or name_lower in s["last"]:
            results.append({k: v for k, v in s.items() if k in ("display_name", "email", "phone", "location", "employee_id")})
    
    return {"results": results[:10], "count": len(results)}


def staff_by_location(location: str):
    """Find all staff at a location."""
    staff = _load_staff()
    if isinstance(staff, dict) and "error" in staff:
        return staff
    
    loc_lower = location.lower().strip()
    results = [s for s in staff if loc_lower in s["location"].lower()]
    
    return {"location": location, "staff": [s["display_name"] for s in results], "count": len(results)}


# ============================================================================
# UNIFIED DASHBOARD — All data for a verified caller, merged by role
# ============================================================================

def get_person_dashboard(name: str, role: str = "staff", email: str = "") -> dict:
    """Return ALL relevant data for a verified caller in one shot.
    
    This merges: QGenda (daily assignments), call schedule, GME balance,
    evaluations, deadlines, and staff roster — all keyed to the person's role.
    """
    result = {
        "person": name,
        "role": role,
        "items": [],
    }
    name_lower = name.lower().strip()
    name_words = name_lower.split()
    first_name = name_words[0] if name_words else ""
    last_name = name_words[-1] if len(name_words) > 1 else ""
    
    # 1. QGenda — today's clinic assignments (everyone)
    try:
        qgenda = _load_qgenda()
        if not isinstance(qgenda, dict) or "error" not in qgenda:
            today = date.today()
            qfmt = f"{today.month:02d}-{today.day:02d}-{str(today.year)[-2:]}"
            qgenda_results = []
            for key in [name_lower, first_name.lower(), last_name.lower()]:
                for e in qgenda["by_person"].get(key, []):
                    if e["date"] == qfmt:
                        qgenda_results.append(e)
            if qgenda_results:
                result["items"].append({
                    "type": "qgenda_today",
                    "label": "Today's Clinic Assignments",
                    "data": {"assignments": qgenda_results, "count": len(qgenda_results)},
                })
            
            # Also get upcoming 7 days
            upcoming = []
            end = today + timedelta(days=7)
            for key in [name_lower, first_name.lower(), last_name.lower()]:
                for e in qgenda["by_person"].get(key, []):
                    parts = e["date"].split("-")
                    if len(parts) == 3:
                        try:
                            d = date(2000 + int(parts[2]), int(parts[0]), int(parts[1]))
                            if today < d <= end:
                                upcoming.append({"date": e["date"], "day": d.strftime("%A"), "task": e["task"]})
                        except:
                            pass
            if upcoming:
                upcoming.sort(key=lambda x: x["date"])
                result["items"].append({
                    "type": "qgenda_upcoming",
                    "label": "Upcoming 7-Day Assignments",
                    "data": {"assignments": upcoming[:20], "total": len(upcoming)},
                })
    except Exception:
        pass
    
    # 2. Call schedule — attending call coverage (faculty only)
    if role in ("administrator", "faculty"):
        try:
            sched = _load_schedule()
            today_str = date.today().strftime("%Y-%m-%d")
            cutoff = (date.today() + timedelta(days=30)).strftime("%Y-%m-%d")
            call_results = []
            for campus, rows in sched.items():
                for row in rows:
                    if row["date"] < today_str or row["date"] > cutoff:
                        continue
                    for role_key in ["primary_clean", "backup_clean", "peds_clean"]:
                        val = row[role_key]
                        if val and name_lower in val.lower():
                            call_results.append({
                                "campus": campus.title(), "date": row["date"],
                                "day": row["day"], "role": role_key.replace("_clean", ""),
                            })
                            break
            if call_results:
                call_results.sort(key=lambda x: x["date"])
                result["items"].append({
                    "type": "call_schedule",
                    "label": "Upcoming Call Coverage",
                    "data": {"assignments": call_results[:10], "total": len(call_results)},
                })
        except Exception:
            pass
    
    # 3. GME balance (residents only)
    if role in ("administrator", "resident"):
        try:
            gme = gme_balance(name)
            if "error" not in gme:
                result["items"].append({
                    "type": "gme_balance",
                    "label": "GME Reimbursement",
                    "data": gme,
                })
        except Exception:
            pass
    
    # 4. Vacation/sick requests — from roster data
    try:
        from modules.roster_parser import query_vacation
        vac = query_vacation(employee_name=name)
        if vac and isinstance(vac, dict) and vac.get("count", 0) > 0:
            result["items"].append({
                "type": "vacation_sick",
                "label": "Vacation & Time Off",
                "data": vac,
            })
    except (ImportError, Exception):
        pass
    
    # 5. Pending evaluations (residents)
    if role in ("administrator", "resident"):
        try:
            from modules import vapi_concierge
            evals = vapi_concierge.get_evaluations_due(name)
            if isinstance(evals, dict) and evals.get("count", 0) > 0:
                result["items"].append({
                    "type": "evaluations",
                    "label": "Evaluations Due",
                    "data": evals,
                })
        except Exception:
            pass
        
        try:
            from modules import vapi_concierge
            deadlines = vapi_concierge.get_deadlines(role)
            if isinstance(deadlines, dict) and deadlines.get("count", 0) > 0:
                result["items"].append({
                    "type": "deadlines",
                    "label": "Program Deadlines",
                    "data": deadlines,
                })
        except Exception:
            pass
    
    return result


# ============================================================================
# LIVE UNIFIED-PLATFORM ROUTING (ADR 0001) — added 2026-07-18
# ============================================================================
# The functions above read flat-file snapshots (xlsx/CSV). The unified
# platform now exposes structured machine-to-machine intents at
# POST /api/v1/vapi/m2m/query (service-token guarded) backed by the LIVE
# databases: urology_qgenda (editable schedule), public.call_schedule,
# CRM contacts, SCL, and reimbursement Postgres.
#
# The wrappers below REBIND the public functions to try LIVE first and fall
# back to the original flat-file implementations on any error, so the voice
# assistant keeps working even if the unified platform is briefly down.
# Live results carry "source": "live" for observability in call logs.
#
# Env (agentic-os container):
#   UNIFIED_API_URL      default http://unified-platform:8097
#   VAPI_SERVICE_TOKEN   shared secret; must match unified-platform .env
#   UNIFIED_API_TIMEOUT  seconds, default 5

import sys as _sys
import urllib.request as _urlreq

UNIFIED_API_URL = os.environ.get("UNIFIED_API_URL", "http://unified-platform:8097").rstrip("/")
UNIFIED_API_TIMEOUT = float(os.environ.get("UNIFIED_API_TIMEOUT", "5"))


def _unified_query(intent: str, params: dict = None) -> dict:
    """POST a structured intent to the unified platform m2m endpoint."""
    token = os.environ.get("VAPI_SERVICE_TOKEN", "")
    if not token:
        raise RuntimeError("VAPI_SERVICE_TOKEN not set")
    body = json.dumps({"intent": intent, "params": params or {}}).encode("utf-8")
    req = _urlreq.Request(
        f"{UNIFIED_API_URL}/api/v1/vapi/m2m/query",
        data=body,
        headers={"Content-Type": "application/json", "X-Service-Token": token},
        method="POST",
    )
    with _urlreq.urlopen(req, timeout=UNIFIED_API_TIMEOUT) as resp:
        payload = json.loads(resp.read().decode("utf-8"))
    data = payload.get("data")
    if data is None:
        raise RuntimeError(f"m2m response missing data for intent {intent}")
    return data


def _live_fail(fn_name: str, err: Exception) -> None:
    print(f"[vapi_unified] live {fn_name} failed, using file fallback: {err}", file=_sys.stderr)


def _strip_t(v) -> str:
    """PG date columns arrive as '2026-07-01T00:00:00.000Z' — keep date part."""
    return str(v or "").split("T")[0]


def _iso_day_name(iso: str) -> str:
    try:
        return datetime.strptime(iso, "%Y-%m-%d").strftime("%A")
    except Exception:
        return ""


def _live_task_label(row: dict) -> str:
    typ = str(row.get("type_name") or row.get("type_category") or "").strip()
    loc = str(row.get("location_name") or "").strip()
    if typ and loc:
        return f"{typ} @ {loc}"
    return typ or loc or "Assigned"


# ── QGenda (urology_qgenda DB — the schedule the unified editor writes) ──

_file_qgenda_today = qgenda_today
_file_qgenda_person_day = qgenda_person_day
_file_qgenda_today_task = qgenda_today_task
_file_qgenda_person_upcoming = qgenda_person_upcoming


def qgenda_today(name=None):  # noqa: F811 — deliberate live rebind
    """LIVE: today's full schedule. If name provided, filter for that person."""
    if name:
        return qgenda_person_day(name)
    try:
        data = _unified_query("today", {})
        entries = [
            {"name": r.get("name", ""), "task": _live_task_label(r), "date": _strip_t(r.get("date"))}
            for r in data.get("assignments", [])
        ]
        if not entries:
            return {"date": data.get("date", ""), "note": "No schedule entries for today",
                    "entries": [], "source": "live"}
        by_task = defaultdict(list)
        for e in entries:
            by_task[e["task"]].append(e["name"])
        return {
            "date": data.get("date", ""),
            "total": len(entries),
            "summary": {t: {"count": len(ns), "people": ns[:10]}
                        for t, ns in sorted(by_task.items(), key=lambda x: -len(x[1]))},
            "entries": entries[:30],
            "source": "live",
        }
    except Exception as e:
        _live_fail("qgenda_today", e)
        return _file_qgenda_today(name)


def qgenda_person_day(name: str, date_str: str = None):  # noqa: F811
    """LIVE: a person's assignments on a date (default today)."""
    try:
        iso = _normalize_date(date_str) if date_str else None
        params = {"name": name}
        if iso:
            params["date"] = iso
        data = _unified_query("person_day", params)
        assignments = [
            {"date": _strip_t(r.get("date")), "task": _live_task_label(r),
             "period": r.get("period") or ""}
            for r in data.get("assignments", [])
        ]
        return {
            "name": data.get("matched") or name,
            "date": _strip_t(data.get("date")) or (iso or date.today().strftime("%Y-%m-%d")),
            "assignments": assignments,
            "count": len(assignments),
            "source": "live",
        }
    except Exception as e:
        _live_fail("qgenda_person_day", e)
        return _file_qgenda_person_day(name, date_str)


def qgenda_today_task(task_name: str):  # noqa: F811
    """LIVE: everyone on a task/clinic/location today."""
    try:
        data = _unified_query("today_task", {"task": task_name})
        people = sorted({str(p.get("name") or "") for p in data.get("people", []) if p.get("name")})
        return {"task": task_name, "date": _strip_t(data.get("date")),
                "people": people, "count": len(people), "source": "live"}
    except Exception as e:
        _live_fail("qgenda_today_task", e)
        return _file_qgenda_today_task(task_name)


def qgenda_person_upcoming(name: str, days: int = 7):  # noqa: F811
    """LIVE: a person's upcoming schedule for the next N days."""
    try:
        data = _unified_query("person_upcoming", {"name": name, "days": int(days)})
        assignments = []
        for r in data.get("assignments", []):
            iso = _strip_t(r.get("date"))
            assignments.append({"date": iso, "day": _iso_day_name(iso), "task": _live_task_label(r)})
        assignments.sort(key=lambda x: x["date"])
        return {"name": data.get("matched") or name, "days": days,
                "assignments": assignments, "count": len(assignments), "source": "live"}
    except Exception as e:
        _live_fail("qgenda_person_upcoming", e)
        return _file_qgenda_person_upcoming(name, days)


# ── Role-based call schedule (postgres public.call_schedule) ──

_file_schedule_by_date = schedule_by_date
_file_call_weekend = call_weekend
_file_call_month = call_month

_CALL_ROLE_COLS = [
    ("primary_attending", "primary"),
    ("backup_attending", "backup"),
    ("peds_attending", "peds"),
    ("chief_resident", "chief"),
    ("first_call_resident", "first_call"),
    ("second_call_resident", "second_call"),
]


def _monday_holiday(mon) -> bool:
    """True if the given Monday is a hospital-observed holiday.

    Covered: New Year's, MLK (3rd Mon Jan), Memorial (last Mon May),
    Juneteenth, July 4th, Labor Day (1st Mon Sep), Christmas —
    including Sunday holidays observed on Monday.
    """
    if mon.weekday() != 0:  # defensive: this rule only applies to Mondays
        return False
    m, d = mon.month, mon.day
    if m == 1 and 15 <= d <= 21:   # MLK — 3rd Monday of January
        return True
    if m == 5 and d >= 25:         # Memorial Day — last Monday of May
        return True
    if m == 9 and d <= 7:          # Labor Day — 1st Monday of September
        return True
    fixed = {(1, 1), (6, 19), (7, 4), (12, 25)}
    if (m, d) in fixed:            # fixed-date holiday falls on Monday
        return True
    prev = mon - timedelta(days=1)
    if (prev.month, prev.day) in fixed:  # fell on Sunday → observed Monday
        return True
    return False


def _weekend_dates() -> list:
    """Weekend block = Fri, Sat, Sun. Monday joins ONLY if it's a holiday.

    Anchored to the current weekend when today is Fri-Sun, otherwise the
    upcoming one.
    """
    today_d = date.today()
    wd = today_d.weekday()
    if wd < 4:
        fri = today_d + timedelta(days=4 - wd)
    else:
        fri = today_d - timedelta(days=wd - 4)
    days = [fri, fri + timedelta(days=1), fri + timedelta(days=2)]
    monday = fri + timedelta(days=3)
    if _monday_holiday(monday):
        days.append(monday)
    return [d.strftime("%Y-%m-%d") for d in days]


def _call_row_shape(row: dict) -> dict:
    iso = _strip_t(row.get("date"))
    return {
        "date": iso,
        "day": row.get("day") or _iso_day_name(iso),
        "primary": row.get("primary_attending") or "—",
        "backup": row.get("backup_attending") or "—",
        "peds": row.get("peds_attending") or "—",
        "chief": row.get("chief_resident") or "—",
        "first_call": row.get("first_call_resident") or "—",
        "second_call": row.get("second_call_resident") or "—",
    }


def schedule_by_date(date_str: str) -> dict:  # noqa: F811
    """LIVE: call coverage (attendings + residents) for a date, all campuses."""
    normalized = _normalize_date(date_str)
    if normalized is None:
        return {
            "date": date_str,
            "note": (
                f"I couldn't understand the date '{date_str}'. "
                "Please say it as a month and day, like 'July 2' or 'July 2 2026', "
                "or use YYYY-MM-DD format."
            ),
        }
    try:
        data = _unified_query("call_schedule", {"date": normalized})
        campuses = {}
        for row in data.get("entries", []):
            hosp = str(row.get("hospital") or "Unknown").title()
            shaped = _call_row_shape(row)
            shaped.pop("date", None)
            shaped.pop("day", None)
            campuses[hosp] = shaped
        if campuses:
            return {"date": normalized, "campuses": campuses, "source": "live"}
        # DB has no rows for this date — the xlsx snapshot may still cover it
        return _file_schedule_by_date(date_str)
    except Exception as e:
        _live_fail("schedule_by_date", e)
        return _file_schedule_by_date(date_str)


def call_weekend():  # noqa: F811
    """LIVE: weekend call block (Fri-Sun; Monday only when it's a holiday)."""
    try:
        dates = _weekend_dates()
        data = _unified_query("call_schedule", {"start": dates[0], "end": dates[-1]})
        entries = data.get("entries", [])
        if not entries:
            return _file_call_weekend()
        result = {"dates": dates, "source": "live"}
        for row in entries:
            hosp = str(row.get("hospital") or "unknown").lower()
            result.setdefault(hosp, []).append(_call_row_shape(row))
        return result
    except Exception as e:
        _live_fail("call_weekend", e)
        return _file_call_weekend()


def call_month(name: str):  # noqa: F811
    """LIVE: a person's call assignments over the next 60 days."""
    try:
        nl = name.lower().strip()
        start = date.today().strftime("%Y-%m-%d")
        end = (date.today() + timedelta(days=60)).strftime("%Y-%m-%d")
        data = _unified_query("call_schedule", {"start": start, "end": end})
        entries = data.get("entries", [])
        if not entries:
            return _file_call_month(name)
        results = []
        for row in entries:
            iso = _strip_t(row.get("date"))
            for col, role in _CALL_ROLE_COLS:
                val = str(row.get(col) or "")
                if val and nl in val.lower():
                    results.append({
                        "campus": str(row.get("hospital") or "").title(),
                        "date": iso,
                        "day": row.get("day") or _iso_day_name(iso),
                        "role": role,
                    })
                    break
        return {"results": results, "total": len(results), "source": "live"}
    except Exception as e:
        _live_fail("call_month", e)
        return _file_call_month(name)


# ── Staff directory (CRM contacts — the source of truth) ──

_file_staff_lookup = staff_lookup
_file_staff_by_location = staff_by_location


def staff_lookup(name: str):  # noqa: F811
    """LIVE: find a staff member in the CRM (source of truth)."""
    try:
        data = _unified_query("staff_lookup", {"name": name})
        results = []
        for c in data.get("results", []):
            results.append({
                "display_name": f"{c.get('first_name', '')} {c.get('last_name', '')}".strip(),
                "email": c.get("email") or "",
                "phone": c.get("mobile") or "",
                "location": c.get("location") or "",
                "category": c.get("category") or "",
                "title": c.get("title") or "",
                "pgy": c.get("pgy") or "",
            })
        return {"results": results[:10], "count": len(results), "source": "live"}
    except Exception as e:
        _live_fail("staff_lookup", e)
        return _file_staff_lookup(name)


def staff_by_location(location: str):  # noqa: F811
    """LIVE: all CRM staff at a location."""
    try:
        data = _unified_query("staff_by_location", {"location": location})
        names = [
            f"{c.get('first_name', '')} {c.get('last_name', '')}".strip()
            for c in data.get("results", [])
        ]
        return {"location": location, "staff": names, "count": len(names), "source": "live"}
    except Exception as e:
        _live_fail("staff_by_location", e)
        return _file_staff_by_location(location)


# ── GME balance — tracker xlsx remains authoritative for spend-vs-cap; ──
# ── live reimbursement system data is attached for context/fallback.   ──

_file_gme_balance = gme_balance


def gme_balance(name: str):  # noqa: F811
    """Tracker xlsx first (authoritative cap math); live reimb data attached."""
    result = _file_gme_balance(name)
    try:
        live = _unified_query("reimb_summary", {"name": name})
        if isinstance(result, dict) and "error" not in result:
            result["live_reimbursement"] = live
        elif isinstance(live, dict) and live.get("found"):
            return {
                "name": live.get("matched") or name,
                "live_reimbursement": live,
                "note": "GME tracker file unavailable; showing live reimbursement system data.",
                "source": "live",
            }
    except Exception:
        pass
    return result
