#!/usr/bin/env python3
"""
Import resident call schedule from RESIDENT_FORM.xlsx into the system.

This script reads the comprehensive resident schedule Excel (which has
Chief Resident, 1st Call Resident, 2nd Call Resident columns alongside
attending data), and produces:

1. data/oncall_schedule_with_residents.json — full canonical JSON with resident fields
2. ~/.hermes/call_schedule_faculty.json — compat format for the backend API
"""
import json
import os
import shutil
from pathlib import Path
from datetime import datetime

import openpyxl

BASE_DIR = Path(__file__).parent
DATA_DIR = BASE_DIR / "data"
RESIDENT_XLSX = BASE_DIR.parent / "Call_Schedule_Q3_Q4_2026_RESIDENT_FORM.xlsx"
CANONICAL_OUT = DATA_DIR / "oncall_schedule_with_residents.json"
COMPAT_OUT = Path.home() / ".hermes" / "call_schedule_faculty.json"

# Also update the existing oncall_schedule.json in place
CANONICAL_EXISTING = DATA_DIR / "oncall_schedule.json"

HOSPITALS = ["Moses", "Wakefield", "Weiler"]

def load_existing_canonical():
    """Load existing oncall_schedule.json for metadata (stats, source)."""
    if CANONICAL_EXISTING.exists():
        try:
            return json.loads(CANONICAL_EXISTING.read_text())
        except Exception as e:
            print(f"Warning: could not load {CANONICAL_EXISTING}: {e}")
    return {"source": RESIDENT_XLSX.name, "entries": [], "hospitals": []}


def parse_resident_xlsx():
    """Parse the RESIDENT_FORM.xlsx and return structured data."""
    wb = openpyxl.load_workbook(RESIDENT_XLSX, data_only=True)
    sheets_data = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        entries = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            raw_date = row[0]
            day = row[1]
            chief = row[2]
            first_res = row[3]
            second_res = row[4]
            primary = row[5]
            backup = row[6]
            peds = row[7]

            if raw_date is None:
                continue

            # Convert datetime or string to ISO date
            if isinstance(raw_date, datetime):
                date_str = raw_date.strftime("%Y-%m-%d")
            else:
                date_str = str(raw_date)

            entries.append({
                "date": date_str,
                "day": day,
                "primary_attending": primary,
                "backup_attending": backup,
                "peds_attending": peds,
                "chief_resident": chief,
                "first_call_resident": first_res,
                "second_call_resident": second_res,
            })

        sheets_data[sheet_name] = {
            "hospital": sheet_name,
            "entries": entries
        }

    return sheets_data


def build_canonical(sheets_data):
    """Build the flat canonical entry list from sheet data."""
    # Merge with existing to preserve any entries beyond the resident form's range
    existing = load_existing_canonical()
    existing_entries = {}
    for e in existing.get("entries", []):
        key = (e["hospital"], e["date"])
        existing_entries[key] = e

    new_entries = []
    seen_keys = set()

    for sheet_name, sheet in sheets_data.items():
        for entry in sheet["entries"]:
            key = (sheet_name, entry["date"])
            seen_keys.add(key)
            new_entries.append({
                "hospital": sheet_name,
                "date": entry["date"],
                "day": entry["day"],
                "primary_attending": entry["primary_attending"],
                "backup_attending": entry["backup_attending"],
                "peds_attending": entry["peds_attending"],
                "chief_resident": entry["chief_resident"],
                "first_call_resident": entry["first_call_resident"],
                "second_call_resident": entry["second_call_resident"],
            })

    # Add any existing entries not in the resident form (e.g. Aug-Dec dates)
    for key, entry in existing_entries.items():
        if key not in seen_keys:
            # Preserve existing fields, add blank resident fields
            new_entry = {
                "hospital": entry.get("hospital", key[0]),
                "date": entry.get("date", key[1]),
                "day": entry.get("day"),
                "primary_attending": entry.get("primary_attending"),
                "backup_attending": entry.get("backup_attending"),
                "peds_attending": entry.get("peds_attending"),
                "chief_resident": entry.get("chief_resident"),
                "first_call_resident": entry.get("first_call_resident"),
                "second_call_resident": entry.get("second_call_resident"),
            }
            new_entries.append(new_entry)

    # Sort by date
    new_entries.sort(key=lambda e: (e["hospital"], e["date"]))

    # Build stats
    hospitals = {}
    for entry in new_entries:
        h = entry["hospital"]
        if h not in hospitals:
            hospitals[h] = {"dates": set(), "primary_docs": set()}
        hospitals[h]["dates"].add(entry["date"])
        if entry["primary_attending"]:
            hospitals[h]["primary_docs"].add(entry["primary_attending"])

    stats = {}
    for h, data in hospitals.items():
        dates = sorted(data["dates"])
        stats[h] = {
            "total_dates": len(dates),
            "start_date": dates[0] if dates else None,
            "end_date": dates[-1] if dates else None,
            "unique_primary_attendings": sorted(data["primary_docs"]),
            "count_unique_attendings": len(data["primary_docs"]),
        }

    canonical = {
        "source": str(RESIDENT_XLSX.name),
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "hospitals": list(hospitals.keys()),
        "stats": stats,
        "entries": new_entries,
    }
    return canonical


def build_compat(canonical):
    """Build the compat JSON format used by the backend API."""
    sheets = {}
    for hospital in HOSPITALS:
        entries = [e for e in canonical["entries"] if e["hospital"] == hospital]
        sheet_entries = []
        for e in entries:
            sheet_entries.append({
                "date": e["date"],
                "day": e["day"],
                "primary": e["primary_attending"],
                "backup": e["backup_attending"],
                "peds": e["peds_attending"],
                "chief_resident": e["chief_resident"],
                "first_call_resident": e["first_call_resident"],
                "second_call_resident": e["second_call_resident"],
            })
        sheets[hospital] = {
            "hospital": hospital,
            "entries": sheet_entries,
        }
    return {"sheets": sheets}


def main():
    print(f"Reading resident schedule from: {RESIDENT_XLSX}")
    sheets_data = parse_resident_xlsx()

    print(f"Building canonical JSON...")
    canonical = build_canonical(sheets_data)

    # Save canonical
    CANONICAL_OUT.parent.mkdir(parents=True, exist_ok=True)
    CANONICAL_OUT.write_text(json.dumps(canonical, indent=2, default=str))
    print(f"Saved canonical: {CANONICAL_OUT} ({len(canonical['entries'])} entries)")

    # Also overwrite existing oncall_schedule.json to keep it in sync
    CANONICAL_EXISTING.write_text(json.dumps(canonical, indent=2, default=str))
    print(f"Updated: {CANONICAL_EXISTING}")

    # Build and save compat
    compat = build_compat(canonical)
    COMPAT_OUT.parent.mkdir(parents=True, exist_ok=True)
    COMPAT_OUT.write_text(json.dumps(compat, indent=2, default=str))
    print(f"Saved compat: {COMPAT_OUT} ({sum(len(s['entries']) for s in compat['sheets'].values())} entries)")

    # Stats
    for hospital in HOSPITALS:
        h_entries = [e for e in canonical["entries"] if e["hospital"] == hospital]
        with_residents = [e for e in h_entries if e.get("chief_resident") or e.get("first_call_resident") or e.get("second_call_resident")]
        print(f"\n{hospital}:")
        print(f"  Total entries: {len(h_entries)}")
        print(f"  With resident data: {len(with_residents)}")
        if with_residents:
            print(f"  Resident date range: {with_residents[0]['date']} to {with_residents[-1]['date']}")

    # Verify some key dates
    print("\n--- Verification: July 6-12 (Moses) ---")
    for entry in canonical["entries"]:
        if entry["hospital"] == "Moses" and "2026-07-06" <= entry["date"] <= "2026-07-12":
            print(f"  {entry['date']} {entry['day']}: Chief={entry['chief_resident']}, 1stCall={entry['first_call_resident']}, 2ndCall={entry['second_call_resident']}, Attending={entry['primary_attending']}")

    print("\nDone!")


if __name__ == "__main__":
    main()
