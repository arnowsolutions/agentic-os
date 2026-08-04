#!/usr/bin/env python3
"""Daily sync: parse xlsx files and update local SQLite cache.
   Reads: Call Schedule, QGenda CSV, Staff CSV, GME Tracker
   Writes to: local SQLite at ~/.hermes/data/sync_cache.db
   Runs via cron once a day.
"""
import csv
import hashlib
import os
import sqlite3
from collections import defaultdict
from datetime import date, datetime

# Data directories - look in /opt/data first, then legacy /workspace
# (Jun 2026 regression: rewrite pointed ONLY at /opt/data, which never existed
#  on any host; the real source files live under /workspace. Fallback restored.)
HERMES_DATA = os.path.join(os.path.dirname(os.path.expanduser("~/.hermes")), "data")
if not os.path.exists(HERMES_DATA):
    os.makedirs(HERMES_DATA, exist_ok=True)
# Canonical cache location consumed by agentic-os (see skill vapi-voice-crm-calendar-swaps)
DB_PATH = "/workspace/agentic-os/data/sync_cache.db"
DATA_BASES = ["/opt/data", "/workspace"]


def _resolve(*rel_parts):
    """Return first existing candidate across DATA_BASES, else the /opt/data path
    (used only to build a meaningful 'file not found' message)."""
    rel = os.path.join(*rel_parts)
    for base in DATA_BASES:
        candidate = os.path.join(base, rel)
        if os.path.exists(candidate):
            return candidate
    return os.path.join(DATA_BASES[0], rel)


QG_PATH = _resolve("repos/qgenda/data/Montefiore_Medical_Center_-_Urology_Schedule_Export_1-1-2026_to_12-31-2026.csv")
STAFF_PATH = _resolve("repos/sick-call-line/data/associates.csv")
SCHED_PATH = _resolve("Call_Schedule_Q3_Q4_2026.xlsx")
GME_PATH = _resolve("Resident_Trackers2025-2026.xlsx")
CHANGES = []


def log(msg):
    print(f"  {msg}")


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS sync_meta (
            table_name TEXT PRIMARY KEY,
            row_count INTEGER,
            checksum TEXT,
            synced_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS qgenda_schedule (
            first_name TEXT, last_name TEXT, email TEXT,
            date TEXT, task TEXT,
            UNIQUE(first_name, last_name, date, task)
        );
        CREATE TABLE IF NOT EXISTS staff_directory (
            display_name TEXT, email TEXT, phone TEXT,
            location TEXT, employee_id TEXT,
            UNIQUE(employee_id)
        );
        CREATE TABLE IF NOT EXISTS call_schedule (
            campus TEXT, date TEXT, day TEXT,
            role TEXT, name TEXT,
            UNIQUE(campus, date, role)
        );
        CREATE TABLE IF NOT EXISTS gme_residents (
            name TEXT, total_spent REAL, remaining REAL,
            cap REAL DEFAULT 1250.0
        );
    """)
    conn.commit()
    conn.close()


def sync_qgenda():
    if not os.path.exists(QG_PATH):
        log(f"SKIP QGenda - file not found: {QG_PATH}")
        return 0
    conn = get_db()
    conn.execute("DELETE FROM qgenda_schedule")
    count = 0
    with open(QG_PATH, encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            conn.execute(
                "INSERT OR IGNORE INTO qgenda_schedule (first_name, last_name, email, date, task) VALUES (?,?,?,?,?)",
                (row.get("Staff First Name","").strip(), row.get("Staff Last Name","").strip(),
                 row.get("Staff Email","").strip(), row.get("Schedule Date","").strip(),
                 row.get("Task Name","").strip())
            )
            count += 1
    conn.commit()
    conn.execute("INSERT OR REPLACE INTO sync_meta (table_name, row_count, checksum, synced_at) VALUES (?,?,?,?)",
                 ("qgenda_schedule", count, str(count), datetime.now().isoformat()))
    conn.commit()
    conn.close()
    CHANGES.append(f"QGenda: {count} rows")
    return count


def sync_staff():
    if not os.path.exists(STAFF_PATH):
        log(f"SKIP Staff - file not found: {STAFF_PATH}")
        return 0
    conn = get_db()
    conn.execute("DELETE FROM staff_directory")
    count = 0
    with open(STAFF_PATH, encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            name = row.get("Employee Name","").strip()
            parts = name.split(",")
            if len(parts) >= 2:
                display = f"{parts[1].strip().split()[0]} {parts[0].strip()}"
            else:
                display = name
            conn.execute(
                "INSERT OR IGNORE INTO staff_directory (display_name, email, phone, location, employee_id) VALUES (?,?,?,?,?)",
                (display, row.get("Email","").strip(), row.get("Phone","").strip(),
                 row.get("Location Code","").strip(), row.get("Employee ID","").strip())
            )
            count += 1
    conn.commit()
    conn.execute("INSERT OR REPLACE INTO sync_meta (table_name, row_count, checksum, synced_at) VALUES (?,?,?,?)",
                 ("staff_directory", count, str(count), datetime.now().isoformat()))
    conn.commit()
    conn.close()
    CHANGES.append(f"Staff: {count} people")
    return count


def sync_call_schedule():
    if not os.path.exists(SCHED_PATH):
        log(f"SKIP Call Schedule - file not found: {SCHED_PATH}")
        return 0
    import openpyxl
    conn = get_db()
    conn.execute("DELETE FROM call_schedule")
    wb = openpyxl.load_workbook(SCHED_PATH, data_only=True)
    count = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for r in range(2, ws.max_row + 1):
            dv = ws.cell(r, 1).value
            if isinstance(dv, datetime):
                ds = dv.strftime("%Y-%m-%d")
            elif isinstance(dv, date):
                ds = dv.strftime("%Y-%m-%d")
            else:
                ds = str(dv or "").strip()[:10]
            day = str(ws.cell(r, 2).value or "")
            for role, col in [("primary", 3), ("backup", 4), ("peds", 5)]:
                name = str(ws.cell(r, col).value or "").strip()
                if name:
                    conn.execute(
                        "INSERT INTO call_schedule (campus, date, day, role, name) VALUES (?,?,?,?,?)",
                        (sheet_name, ds, day, role, name)
                    )
                    count += 1
    conn.commit()
    conn.execute("INSERT OR REPLACE INTO sync_meta (table_name, row_count, checksum, synced_at) VALUES (?,?,?,?)",
                 ("call_schedule", count, str(count), datetime.now().isoformat()))
    conn.commit()
    conn.close()
    CHANGES.append(f"Call Schedule: {count} assignments")
    return count


def sync_gme():
    if not os.path.exists(GME_PATH):
        log(f"SKIP GME - file not found: {GME_PATH}")
        return 0
    import openpyxl
    conn = get_db()
    conn.execute("DELETE FROM gme_residents")
    wb = openpyxl.load_workbook(GME_PATH, data_only=True)
    ws = wb["Sheet 2025-2026"]
    residents = defaultdict(float)
    for r in range(2, ws.max_row + 1):
        name = str(ws.cell(r, 1).value or "").strip()
        amount = ws.cell(r, 5).value
        if name and amount and isinstance(amount, (int, float)):
            residents[name] += abs(amount)
    count = 0
    for name, spent in residents.items():
        remaining = max(0, 1250.0 - min(spent, 1250.0))
        conn.execute(
            "INSERT INTO gme_residents (name, total_spent, remaining, cap) VALUES (?,?,?,?)",
            (name, round(spent, 2), round(remaining, 2), 1250.0)
        )
        count += 1
    conn.commit()
    conn.execute("INSERT OR REPLACE INTO sync_meta (table_name, row_count, checksum, synced_at) VALUES (?,?,?,?)",
                 ("gme_residents", count, str(count), datetime.now().isoformat()))
    conn.commit()
    conn.close()
    CHANGES.append(f"GME: {count} residents")
    return count


if __name__ == "__main__":
    os.makedirs(HERMES_DATA, exist_ok=True)
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    init_db()
    sync_qgenda()
    sync_staff()
    sync_call_schedule()
    sync_gme()
    print(f"Sync complete at {datetime.now().isoformat()}")
    if CHANGES:
        for c in CHANGES:
            print(f"  {c}")
    else:
        print("  No source data files found - all syncs skipped.")
