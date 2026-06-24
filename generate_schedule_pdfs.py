#!/usr/bin/env python3
"""
Generate Schedule PDFs — Montefiore Urology Call Schedule
Standalone CLI script using reportlab.

Reads the master xlsx for attending data, optionally merges with a
resident-completed xlsx (Chief Resident, 1st Call Resident, 2nd Call Resident),
and generates 3 landscape-letter PDFs (Moses, Wakefield, Weiler).

Usage:
    python generate_schedule_pdfs.py
    python generate_schedule_pdfs.py --master /path/to/master.xlsx
    python generate_schedule_pdfs.py --residents /path/to/residents.xlsx
    python generate_schedule_pdfs.py --output /custom/output/dir
"""

import argparse
import os
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

# ── Constants ──────────────────────────────────────────────────

SHEETS = ["Moses", "Wakefield", "Weiler"]

# Column indices (1-based) in the master xlsx
COL_DATE = 1
COL_DAY = 2
COL_PRIMARY = 3
COL_BACKUP = 4
COL_PEDS = 5
# Resident columns (if present in uploaded file or master)
COL_CHIEF = 8
COL_CALL1 = 6
COL_CALL2 = 7

HEADERS = [
    "Date",
    "Day",
    "Chief Resident",
    "1st Call Resident",
    "2nd Call Resident",
    "Primary Attending",
    "Backup Attending",
    "PEDS Attending",
]

MERGE_BLANK_PLACEHOLDER = "_______________"
DARK_HEADER = colors.HexColor("#1a3a5c")
LIGHT_GRAY = colors.HexColor("#f5f5f5")
WHITE = colors.white
WEEKEND_BLUE = colors.HexColor("#e4edf5")
MONTH_SEPARATOR = colors.HexColor("#2c5f8a")
ACCENT_LINE = colors.HexColor("#3a7bd5")

DEFAULT_MASTER = "/workspace/Call_Schedule_Q3_Q4_2026.xlsx"


# ── Data Loading ───────────────────────────────────────────────

def load_master_data(xlsx_path: str) -> dict:
    """Load attending data from the master xlsx.
    Returns {sheet_name: [{date, day, primary, backup, peds}, ...]}
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    data = {}
    for sheet_name in SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"Warning: Sheet '{sheet_name}' not found in {xlsx_path}, skipping")
            continue
        ws = wb[sheet_name]
        rows = []
        for r in range(2, ws.max_row + 1):
            dt_val = ws.cell(r, COL_DATE).value
            if dt_val is None:
                continue
            if hasattr(dt_val, "strftime"):
                dt = dt_val
            else:
                continue
            rows.append({
                "date": dt,
                "day": str(ws.cell(r, COL_DAY).value or ""),
                "primary": str(ws.cell(r, COL_PRIMARY).value or ""),
                "backup": str(ws.cell(r, COL_BACKUP).value or ""),
                "peds": str(ws.cell(r, COL_PEDS).value or ""),
                "chief": "",
                "call1": "",
                "call2": "",
            })
        data[sheet_name] = rows
    wb.close()
    return data


def load_resident_data(xlsx_path: str) -> dict:
    """Load resident columns from an uploaded xlsx.
    Looks for columns named: Chief Resident, 1st Call Resident, 2nd Call Resident
    (or similar variants).

    Returns {sheet_name: {date_str: {chief, call1, call2}}}
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    result = {}

    for sheet_name in SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        # Read header row to find column indices
        headers = {}
        for col in range(1, ws.max_column + 1):
            val = str(ws.cell(1, col).value or "").strip().lower()
            if val:
                headers[col] = val

        # Map header names to column indices
        chief_col = None
        call1_col = None
        call2_col = None
        date_col = 1

        for col, name in headers.items():
            if "chief" in name:
                chief_col = col
            elif "1st" in name or "first" in name or "call1" in name:
                call1_col = col
            elif "2nd" in name or "second" in name or "call2" in name:
                call2_col = col

        sheet_data = {}
        for r in range(2, ws.max_row + 1):
            dt_val = ws.cell(r, date_col).value
            if dt_val is None:
                continue
            if hasattr(dt_val, "strftime"):
                date_key = dt_val.strftime("%Y-%m-%d")
            elif hasattr(dt_val, "isoformat"):
                date_key = dt_val.isoformat()[:10]
            else:
                continue

            entry = {
                "chief": str(ws.cell(r, chief_col).value or "").strip() if chief_col else "",
                "call1": str(ws.cell(r, call1_col).value or "").strip() if call1_col else "",
                "call2": str(ws.cell(r, call2_col).value or "").strip() if call2_col else "",
            }
            sheet_data[date_key] = entry

        result[sheet_name] = sheet_data

    wb.close()
    return result


def merge_data(master: dict, resident_data: dict = None) -> dict:
    """Merge resident data into master data by date+sheet."""
    merged = {}
    for sheet_name in SHEETS:
        if sheet_name not in master:
            continue
        merged_rows = []
        sheet_residents = (resident_data or {}).get(sheet_name, {})
        for row in master[sheet_name]:
            date_key = row["date"].strftime("%Y-%m-%d")
            resid = sheet_residents.get(date_key, {})
            merged_rows.append({
                "date": row["date"],
                "day": row["day"],
                "primary": row["primary"],
                "backup": row["backup"],
                "peds": row["peds"],
                "chief": resid.get("chief", row.get("chief", "") or "_______________"),
                "call1": resid.get("call1", row.get("call1", "") or "_______________"),
                "call2": resid.get("call2", row.get("call2", "") or "_______________"),
            })
        merged[sheet_name] = merged_rows
    return merged


# ── PDF Generation ─────────────────────────────────────────────

def generate_pdfs(data: dict, output_dir: str) -> list:
    """Generate 3 PDFs (Moses, Wakefield, Weiler) into output_dir.
    Returns list of PDF file paths.
    """
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    pdf_paths = []

    for sheet_name in SHEETS:
        if sheet_name not in data:
            continue
        rows = data[sheet_name]
        if not rows:
            continue

        pdf_path = str(Path(output_dir) / f"{sheet_name.lower()}_call_schedule.pdf")
        _generate_single_pdf(sheet_name, rows, pdf_path)
        pdf_paths.append(pdf_path)
        print(f"  ✓ Generated {pdf_path}")

    return pdf_paths


def _generate_single_pdf(sheet_name: str, rows: list, output_path: str):
    """Generate a single landscape-letter PDF for one sheet using reportlab."""
    doc = SimpleDocTemplate(
        output_path,
        pagesize=landscape(letter),
        leftMargin=0.3 * inch,
        rightMargin=0.3 * inch,
        topMargin=0.3 * inch,
        bottomMargin=0.3 * inch,
    )

    # ── Styles ──
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ScheduleTitle",
        parent=styles["Normal"],
        fontSize=12,
        textColor=DARK_HEADER,
        spaceAfter=2,
        spaceBefore=0,
        fontName="Helvetica-Bold",
    )
    subtitle_style = ParagraphStyle(
        "ScheduleSubtitle",
        parent=styles["Normal"],
        fontSize=8,
        textColor=colors.HexColor("#555555"),
        spaceAfter=6,
        spaceBefore=0,
        fontName="Helvetica",
    )
    month_header_style = ParagraphStyle(
        "MonthHeader",
        parent=styles["Normal"],
        fontSize=9,
        textColor=WHITE,
        spaceAfter=0,
        spaceBefore=0,
        fontName="Helvetica-Bold",
        alignment=TA_CENTER,
    )
    cell_style = ParagraphStyle(
        "Cell",
        parent=styles["Normal"],
        fontSize=6.5,
        textColor=colors.HexColor("#333333"),
        spaceAfter=0,
        spaceBefore=0,
        leading=8,
        fontName="Helvetica",
        alignment=TA_CENTER,
    )
    cell_style_strong = ParagraphStyle(
        "CellStrong",
        parent=styles["Normal"],
        fontSize=6.5,
        textColor=DARK_HEADER,
        spaceAfter=0,
        spaceBefore=0,
        leading=8,
        fontName="Helvetica-Bold",
        alignment=TA_CENTER,
    )
    header_style = ParagraphStyle(
        "HeaderCell",
        parent=styles["Normal"],
        fontSize=7,
        textColor=WHITE,
        spaceAfter=0,
        spaceBefore=0,
        leading=9,
        fontName="Helvetica-Bold",
        alignment=TA_CENTER,
    )
    footer_style = ParagraphStyle(
        "Footer",
        parent=styles["Normal"],
        fontSize=6,
        textColor=colors.HexColor("#999999"),
        spaceAfter=0,
        spaceBefore=4,
        fontName="Helvetica",
    )

    # ── Group rows by month ──
    months = {}  # month_label -> [rows]
    for row in rows:
        month_label = row["date"].strftime("%B %Y")
        months.setdefault(month_label, []).append(row)

    # Sort months chronologically
    sorted_months = sorted(months.keys(), key=lambda m: datetime.strptime(m, "%B %Y"))

    elements = []

    # ── Title block ──
    elements.append(Spacer(1, 2))
    elements.append(Paragraph(
        f"Montefiore Urology — {sheet_name} Call Schedule",
        title_style,
    ))
    elements.append(Paragraph(
        f"Q3–Q4 2026  |  Generated {datetime.now().strftime('%B %d, %Y at %I:%M %p')}",
        subtitle_style,
    ))

    # ── Column widths (8 columns, landscape letter ≈ 10.4in usable width) ──
    usable_width = landscape(letter)[0] - 0.6 * inch  # ~9.4in
    col_widths = [
        usable_width * 0.10,  # Date
        usable_width * 0.07,  # Day
        usable_width * 0.14,  # Chief
        usable_width * 0.14,  # 1st Call
        usable_width * 0.14,  # 2nd Call
        usable_width * 0.14,  # Primary
        usable_width * 0.14,  # Backup
        usable_width * 0.13,  # PEDS
    ]

    # ── Month by month ──
    for month_idx, month_label in enumerate(sorted_months):
        month_rows = months[month_label]

        # Month separator row
        month_sep = [
            Paragraph(f"<b>{month_label}</b>", month_header_style)
        ] * 8
        sep_table = Table([month_sep], colWidths=col_widths)
        sep_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), MONTH_SEPARATOR),
            ("TEXTCOLOR", (0, 0), (-1, -1), WHITE),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("GRID", (0, 0), (-1, -1), 0.5, MONTH_SEPARATOR),
        ]))
        elements.append(sep_table)

        # Header row
        header_cells = [Paragraph(h, header_style) for h in HEADERS]
        header_table = Table([header_cells], colWidths=col_widths)
        header_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), DARK_HEADER),
            ("TEXTCOLOR", (0, 0), (-1, -1), WHITE),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("GRID", (0, 0), (-1, -1), 0.5, DARK_HEADER),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ]))
        elements.append(header_table)

        # Data rows
        for row_idx, row in enumerate(month_rows):
            day_name = row["day"][:3].lower()
            is_weekend = day_name in ("sat", "sun")

            date_str = row["date"].strftime("%b %d")
            day_str = row["day"][:3]

            cells = [
                Paragraph(date_str, cell_style),
                Paragraph(day_str, cell_style),
                Paragraph(row["chief"], cell_style_strong if row["chief"] else cell_style),
                Paragraph(row["call1"], cell_style),
                Paragraph(row["call2"], cell_style),
                Paragraph(row["primary"], cell_style),
                Paragraph(row["backup"], cell_style),
                Paragraph(row["peds"], cell_style),
            ]

            data_table = Table([cells], colWidths=col_widths)
            row_bg = WEEKEND_BLUE if is_weekend else (LIGHT_GRAY if row_idx % 2 == 1 else WHITE)
            data_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), row_bg),
                ("TOPPADDING", (0, 0), (-1, -1), 1.5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 1.5),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#cccccc")),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]))
            elements.append(data_table)

        # Small spacer between months
        elements.append(Spacer(1, 3))

    # ── Footer ──
    elements.append(Spacer(1, 4))
    elements.append(Paragraph(
        f"Montefiore Urology Call Schedule — {sheet_name} — Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        footer_style,
    ))

    # Build PDF
    doc.build(elements)
    return output_path


# ── CLI Entry Point ────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Generate Montefiore Urology Call Schedule PDFs",
    )
    parser.add_argument(
        "--master",
        default=DEFAULT_MASTER,
        help=f"Path to master xlsx (default: {DEFAULT_MASTER})",
    )
    parser.add_argument(
        "--residents",
        default=None,
        help="Path to resident-completed xlsx with Chief/1st Call/2nd Call columns",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Output directory (default: reports/schedules/<timestamp>/)",
    )
    args = parser.parse_args()

    # Validate master file
    master_path = args.master
    if not os.path.exists(master_path):
        print(f"Error: Master file not found: {master_path}", file=sys.stderr)
        sys.exit(1)

    # Resolve output directory
    if args.output:
        output_dir = args.output
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        script_dir = Path(__file__).parent.resolve()
        output_dir = str(script_dir / "reports" / "schedules" / timestamp)

    print(f"📋 Reading master schedule: {master_path}")
    master_data = load_master_data(master_path)
    print(f"   Loaded {sum(len(v) for v in master_data.values())} rows across {len(master_data)} sheets")

    resident_data = None
    if args.residents:
        if not os.path.exists(args.residents):
            print(f"Error: Residents file not found: {args.residents}", file=sys.stderr)
            sys.exit(1)
        print(f"📋 Reading resident data: {args.residents}")
        resident_data = load_resident_data(args.residents)
        print(f"   Loaded resident data for {len(resident_data)} sheets")

    print("🔀 Merging data...")
    merged = merge_data(master_data, resident_data)

    print(f"📄 Generating PDFs → {output_dir}")
    pdfs = generate_pdfs(merged, output_dir)

    print(f"\n✅ Done! {len(pdfs)} PDFs generated:")
    for p in pdfs:
        size = os.path.getsize(p)
        print(f"   • {p} ({size / 1024:.1f} KB)")

    return pdfs


if __name__ == "__main__":
    main()
