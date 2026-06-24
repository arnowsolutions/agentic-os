#!/usr/bin/env python3
"""
Schedule Reporter — reads call/staff schedules from xlsx sources
and emails formatted reports on-demand.

Usage:
  python3 schedule-report.py --call today --email sfrasier@montefiore.org
  python3 schedule-report.py --call week --email jessie@montefiore.org
  python3 schedule-report.py --staff today --email winnie@montefiore.org
"""
import os, sys, json, yaml
from datetime import datetime, date, timedelta, timezone
from pathlib import Path
from collections import defaultdict
import openpyxl
import re

sys.path.insert(0, os.path.expanduser("~/.hermes/home/.local/lib/python3.12/site-packages"))
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

CALL_SCHEDULE_PATH = Path("/workspace/Call_Schedule_Q3_Q4_2026.xlsx")
STAFF_SCHEDULE_DIR = Path("/workspace")
DEFAULT_RECIPIENT = "sfrasier@montefiore.org"

SIG_BLOCK = ""
_CONFIG_PATH = Path.home() / ".hermes" / "email_accounts.yaml"
if _CONFIG_PATH.exists():
    try:
        _config = yaml.safe_load(_CONFIG_PATH.read_text())
        _accts = (_config or {}).get('accounts', {})
        _sigs = [v.get('signature_html', '') for v in _accts.values() if v.get('signature_html')]
        if _sigs:
            SIG_BLOCK = _sigs[0]
    except Exception:
        pass


def get_greeting():
    et = datetime.now(timezone.utc).astimezone()
    h = et.hour
    return "Good Morning" if h < 12 else "Good Afternoon" if h < 17 else "Good Evening"


def get_call_schedule(period="today"):
    """Get call schedule data from the xlsx. Period: today, week, month, or YYYY-MM-DD."""
    if not CALL_SCHEDULE_PATH.exists():
        return None, "Call schedule file not found"

    wb = openpyxl.load_workbook(str(CALL_SCHEDULE_PATH), data_only=True)
    
    now = datetime.now()
    today = now.date()
    
    if period == "today":
        start_date = today
        end_date = today
    elif period == "week":
        start_date = today
        end_date = today + timedelta(days=7)
    elif period == "month":
        start_date = today
        if today.month == 12:
            end_date = date(today.year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(today.year, today.month + 1, 1) - timedelta(days=1)
    else:
        try:
            start_date = datetime.strptime(period, "%Y-%m-%d").date()
            end_date = start_date
        except ValueError:
            return None, f"Invalid period: {period}"
    
    locations = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        location_data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            date_val = row[0]
            if not date_val:
                continue
            row_date = date_val.date() if isinstance(date_val, datetime) else (date_val if isinstance(date_val, date) else datetime.strptime(str(date_val)[:10], "%Y-%m-%d").date())
            
            if start_date <= row_date <= end_date:
                location_data.append({
                    'date': row_date,
                    'day': str(row[1] or '') if len(row) > 1 else '',
                    'primary': str(row[2] or '') if len(row) > 2 else '',
                    'backup': str(row[3] or '') if len(row) > 3 else '',
                    'peds': str(row[4] or '') if len(row) > 4 else '',
                })
        if location_data:
            locations[sheet_name] = location_data
    
    wb.close()
    
    if not locations:
        # Auto-find next available date range
        wb2 = openpyxl.load_workbook(str(CALL_SCHEDULE_PATH), data_only=True)
        all_dates = []
        for sheet_name in wb2.sheetnames:
            ws = wb2[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    d = row[0].date() if isinstance(row[0], datetime) else row[0]
                    all_dates.append(d)
        wb2.close()
        
        if all_dates:
            next_date = min(all_dates)
            if period == "today" or period == "week":
                start_date = next_date
                end_date = next_date + timedelta(days=6)
                # Re-fetch
                wb3 = openpyxl.load_workbook(str(CALL_SCHEDULE_PATH), data_only=True)
                locations = {}
                for sheet_name in wb3.sheetnames:
                    ws = wb3[sheet_name]
                    location_data = []
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        date_val = row[0]
                        if not date_val: continue
                        row_date = date_val.date() if isinstance(date_val, datetime) else date_val
                        if start_date <= row_date <= end_date:
                            location_data.append({
                                'date': row_date,
                                'day': str(row[1] or ''),
                                'primary': str(row[2] or ''),
                                'backup': str(row[3] or ''),
                                'peds': str(row[4] or ''),
                            })
                    if location_data:
                        locations[sheet_name] = location_data
                wb3.close()
                return locations, None
        return None, f"No schedule data found for {start_date} to {end_date}"
    
    return locations, None


def format_call_for_email(locations, period):
    """Build HTML for call schedule email."""
    greeting = get_greeting()
    now = datetime.now()
    
    if period == "today":
        period_label = "Today"
    elif period == "week":
        period_label = "This Week"
    elif period == "month":
        period_label = "This Month"
    else:
        period_label = period
    
    # Group by date
    all_dates = set()
    for loc, entries in locations.items():
        for e in entries:
            all_dates.add(e['date'])
    all_dates = sorted(all_dates)
    
    date_sections = ""
    for d in all_dates:
        date_str = d.strftime("%A, %B %d, %Y")
        
        loc_rows = ""
        for loc_name in sorted(locations.keys()):
            entries = locations[loc_name]
            for e in entries:
                if e['date'] == d:
                    backup_html = f"<br><span style='font-size:8pt;color:#888'>Backup: {e['backup']}</span>" if e['backup'] and e['backup'] != 'None' else ""
                    peds_html = f"<br><span style='font-size:8pt;color:#888'>PEDS: {e['peds']}</span>" if e['peds'] and e['peds'] != 'None' else ""
                    loc_rows += f"""
            <tr>
              <td style="padding:6px 10px;border-bottom:1px solid #e8e8e8;font-family:Georgia,'Times New Roman',serif;font-size:10pt;color:#555;width:100px;font-weight:bold">{loc_name}</td>
              <td style="padding:6px 10px;border-bottom:1px solid #e8e8e8;font-family:Georgia,'Times New Roman',serif;font-size:10pt;color:#333">
                {e['primary']}{backup_html}{peds_html}
              </td>
            </tr>"""
        
        date_sections += f"""
      <tr>
        <td style="padding:12px 0 4px 0;font-family:Georgia,'Times New Roman',serif;font-size:12pt;font-weight:bold;color:#1a3a5c">{date_str}</td>
      </tr>
      <tr>
        <td style="padding:0">
          <table cellpadding="0" cellspacing="0" width="100%" style="border-collapse:collapse">
            <tr style="color:#888;font-size:8pt;font-family:Georgia,'Times New Roman',serif">
              <td style="padding:4px 10px;border-bottom:2px solid #d0d4d8;width:100px;font-weight:bold">Location</td>
              <td style="padding:4px 10px;border-bottom:2px solid #d0d4d8;font-weight:bold">Attending</td>
            </tr>
            {loc_rows}
          </table>
        </td>
      </tr>"""
    
    html = f"""<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0"></head>
<body style="margin:0;padding:0;background-color:#f0f2f5;font-family:'Times New Roman',Georgia,serif">
<table cellpadding="0" cellspacing="0" width="100%" style="background-color:#f0f2f5">
  <tr><td style="padding:30px 10px" align="center">
    <table cellpadding="0" cellspacing="0" style="max-width:620px;width:100%;background-color:#ffffff;border-radius:4px;box-shadow:0 1px 3px rgba(0,0,0,0.08)">
      <tr>
        <td style="background-color:#1a3a5c;border-radius:4px 4px 0 0;padding:18px 28px">
          <table cellpadding="0" cellspacing="0" width="100%">
            <tr>
              <td style="font-family:Georgia,'Times New Roman',serif;font-size:16pt;font-weight:bold;color:#ffffff">Montefiore Einstein Urology</td>
              <td style="font-family:Georgia,'Times New Roman',serif;font-size:9pt;color:#b0c4de;text-align:right;vertical-align:bottom">{now.strftime('%B %d, %Y')}</td>
            </tr>
          </table>
        </td>
      </tr>
      <tr>
        <td style="padding:16px 28px 0 28px;font-family:Georgia,'Times New Roman',serif;font-size:16pt;font-weight:bold;color:#1a3a5c">Call Schedule — {period_label}</td>
      </tr>
      <tr><td style="padding:0 28px"><hr style="border:none;border-top:1px solid #d0d4d8;margin:0"></td></tr>
      <tr>
        <td style="padding:14px 28px 0 28px">
          <p style="margin:0 0 4px 0;font-family:Times New Roman,Georgia,serif;font-size:12pt;color:#333;line-height:1.5">{greeting} Shareef,</p>
          <p style="margin:0 0 12px 0;font-family:Times New Roman,Georgia,serif;font-size:11pt;color:#333;line-height:1.5">Here is the call schedule for <b>{period_label.lower()}</b>.</p>
        </td>
      </tr>
      {date_sections}
      <tr>
        <td style="padding:10px 28px 24px 28px">
          <hr style="border:none;border-top:1px solid #ddd;margin:0 0 10px 0">
          {SIG_BLOCK}
        </td>
      </tr>
    </table>
    <table cellpadding="0" cellspacing="0" style="max-width:620px;width:100%;padding-top:10px">
      <tr><td style="font-family:Arial,Helvetica,sans-serif;font-size:8pt;color:#aaa;text-align:center">Montefiore Urology &bull; 1250 Waters Place &bull; Bronx, NY 10461</td></tr>
    </table>
  </td></tr>
</table>
</body>
</html>"""
    return html


def generate_call_pdf(locations, period_label):
    """Generate PDF for call schedule."""
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.lib.colors import HexColor
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    import tempfile
    
    pdf = tempfile.NamedTemporaryFile(suffix='.pdf', delete=False)
    pdf_path = pdf.name
    pdf.close()
    
    doc = SimpleDocTemplate(pdf_path, pagesize=letter,
        topMargin=0.6*inch, bottomMargin=0.6*inch,
        leftMargin=0.6*inch, rightMargin=0.6*inch)
    
    styles = getSampleStyleSheet()
    s_title = ParagraphStyle('Title', parent=styles['Title'],
        textColor=HexColor('#1a3a5c'), fontSize=18, spaceAfter=2, fontName='Helvetica-Bold')
    s_sub = ParagraphStyle('Sub', parent=styles['Normal'],
        textColor=HexColor('#666666'), fontSize=10, spaceAfter=16)
    s_h2 = ParagraphStyle('H2', parent=styles['Heading2'],
        textColor=HexColor('#1a3a5c'), fontSize=14, spaceBefore=12, spaceAfter=6, fontName='Helvetica-Bold')
    s_body = ParagraphStyle('Body', parent=styles['Normal'],
        textColor=HexColor('#333333'), fontSize=9, leading=13, spaceAfter=4)
    s_muted = ParagraphStyle('Muted', parent=s_body, textColor=HexColor('#888888'), fontSize=8)
    s_right = ParagraphStyle('Right', parent=s_body, alignment=TA_RIGHT, fontSize=9)
    
    elements = []
    
    # Header
    hdr = Table([
        [Paragraph("<b>Montefiore Einstein Urology</b>", ParagraphStyle('Hdr', parent=s_title, fontSize=14, alignment=TA_LEFT)),
         Paragraph("Call Schedule", ParagraphStyle('HdrDate', parent=s_sub, fontSize=12, textColor=HexColor('#ffffff'), alignment=TA_RIGHT))]
    ], colWidths=[4.3*inch, 3*inch])
    hdr.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), HexColor('#1a3a5c')),
        ('TEXTCOLOR', (0, 0), (-1, -1), HexColor('#ffffff')),
        ('LEFTPADDING', (0, 0), (0, 0), 14), ('RIGHTPADDING', (-1, -1), (-1, -1), 14),
        ('TOPPADDING', (0, 0), (-1, -1), 10), ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(hdr)
    elements.append(Spacer(1, 12))
    
    now = datetime.now()
    elements.append(Paragraph(f"<b>Call Schedule — {period_label}</b>", s_h2))
    elements.append(Paragraph(f"Generated {now.strftime('%B %d, %Y')}", s_sub))
    
    # Group by date
    all_dates = set()
    for loc, entries in locations.items():
        for e in entries:
            all_dates.add(e['date'])
    all_dates = sorted(all_dates)
    
    for d in all_dates:
        elements.append(Paragraph(f"<b>{d.strftime('%A, %B %d, %Y')}</b>",
            ParagraphStyle('DayHdr', parent=s_body, fontSize=11, spaceBefore=14, spaceAfter=6, fontName='Helvetica-Bold', textColor=HexColor('#1a3a5c'))))
        
        txn_data = [[
            Paragraph("<b>Location</b>", s_muted),
            Paragraph("<b>Primary</b>", s_muted),
            Paragraph("<b>Backup</b>", s_muted),
            Paragraph("<b>PEDS</b>", s_muted),
        ]]
        for loc_name in sorted(locations.keys()):
            for e in locations[loc_name]:
                if e['date'] == d:
                    txn_data.append([
                        Paragraph(loc_name, s_body),
                        Paragraph(e['primary'] if e['primary'] and e['primary'] != 'None' else '—', s_body),
                        Paragraph(e['backup'] if e['backup'] and e['backup'] != 'None' else '—', s_body),
                        Paragraph(e['peds'] if e['peds'] and e['peds'] != 'None' else '—', s_muted),
                    ])
        
        if len(txn_data) > 1:
            t = Table(txn_data, colWidths=[1.0*inch, 2.5*inch, 2.0*inch, 1.5*inch])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), HexColor('#f4f6f9')),
                ('GRID', (0, 0), (-1, -1), 0.25, HexColor('#e0e4e8')),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ]))
            elements.append(t)
    
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(
        "<i>Generated by Hermes Agent — Montefiore Urology</i>",
        ParagraphStyle('Footer', parent=s_body, textColor=HexColor('#999999'), fontSize=7, alignment=TA_CENTER)))
    
    doc.build(elements)
    return pdf_path


def send_call_schedule(period, recipient):
    """Send call schedule email with PDF attachment."""
    from google_workspace import GoogleWorkspace
    
    locations, error = get_call_schedule(period)
    if error:
        return f"❌ {error}"
    
    html = format_call_for_email(locations, period)
    
    now = datetime.now()
    if period == "today":
        period_label = f"Today ({now.strftime('%b %d')})"
    elif period == "week":
        period_label = f"Week of {now.strftime('%b %d')}"
    elif period == "month":
        period_label = f"{now.strftime('%B %Y')}"
    else:
        period_label = period
    
    # Generate PDF
    pdf_path = generate_call_pdf(locations, period_label)
    
    ws = GoogleWorkspace()
    result = ws.send_email(
        user_id="urologyresidency",
        to=recipient,
        subject=f"Montefiore Urology — Call Schedule ({period_label})",
        body=html,
        attachments=[pdf_path],
        from_email="sfrasier@montefiore.org",
        is_html=True,
    )
    
    if result.get("successful"):
        return f"✅ Call schedule sent to {recipient}"
    else:
        return f"❌ Failed: {json.dumps(result, indent=2, default=str)[:300]}"


SHIFT_LEGEND = {
    'HT': 'Hutch Tower', 'MAP': 'MAP', 'WF': 'Wakefield',
    'KSP': 'KSP', 'BA': 'Brook Ave', 'V': 'Vacation',
    'LOA': 'Leave of Absence',
}

def get_staff_schedule(period="today"):
    """Get staff schedule from the latest monthly xlsx. Period: today, week, month."""
    # Find the latest monthly schedule file
    pattern = re.compile(r'Urology 2026 Monthly Schedule \(([A-Z]+)(\d+)-([A-Z]+)(\d+)\)\.xlsx')
    latest_file = None
    latest_end = None
    
    for f in STAFF_SCHEDULE_DIR.glob("Urology 2026 Monthly Schedule*.xlsx"):
        m = pattern.search(f.name)
        if m:
            month_map = {'JAN':1,'FEB':2,'MAR':3,'APR':4,'MAY':5,'JUN':6,'JUL':7,'AUG':8,'SEP':9,'SEPT':9,'OCT':10,'NOV':11,'DEC':12}
            end_month = month_map.get(m.group(3).upper(), 0)
            end_day = int(m.group(4))
            try:
                end_date = date(2026, end_month, end_day)
                if latest_end is None or end_date > latest_end:
                    latest_end = end_date
                    latest_file = f
            except ValueError:
                continue
    
    if not latest_file:
        return None, "No monthly schedule file found"
    
    wb = openpyxl.load_workbook(str(latest_file), data_only=True)
    
    now = datetime.now()
    today = now.date()
    
    if period == "today":
        start_date = today
        end_date = today
    elif period == "week":
        start_date = today
        end_date = today + timedelta(days=7)
    elif period == "month":
        start_date = today
        end_date = today + timedelta(days=28)
    else:
        return None, f"Invalid period: {period}"
    
    results = {}
    for sheet_name in wb.sheetnames:
        if sheet_name == 'WINK':
            continue
        ws = wb[sheet_name]
        
        # Get date headers
        date_cols = {}
        for c in range(3, min(ws.max_column+1, 35)):
            val = ws.cell(9, c).value
            if val:
                d = val.date() if isinstance(val, datetime) else val
                date_cols[d] = c
        
        if not date_cols:
            continue
        
        schedule_dates = sorted(date_cols.keys())
        sheet_start = schedule_dates[0]
        sheet_end = schedule_dates[-1]
        
        # Check if our query range overlaps with this sheet's range
        if start_date > sheet_end or end_date < sheet_start:
            continue
        
        # If period is "today" and today isn't in range, use sheet_start
        effective_start = max(start_date, sheet_start) if start_date >= sheet_start else sheet_start
        effective_end = min(end_date, sheet_end)
        
        if effective_start > effective_end:
            continue
        
        # Parse staff data
        staff_data = []
        for r in range(10, ws.max_row + 1):
            name = ws.cell(r, 1).value
            if not name or not str(name).strip():
                continue
            name_str = str(name).strip()
            if 'TOTAL' in name_str.upper() or 'SUBTOTAL' in name_str.upper():
                continue
            
            role = str(ws.cell(r, 2).value or '').strip()
            # Skip site rows
            if role == 'Site':
                continue
            
            daily = []
            current = effective_start
            while current <= effective_end:
                # Skip weekends (Saturday=5, Sunday=6)
                if current.weekday() >= 5:
                    current += timedelta(days=1)
                    continue
                
                col = date_cols.get(current)
                if col:
                    shift = ws.cell(r, col).value
                    shift_str = str(shift) if shift else ''
                    # Skip OFF staff
                    if shift_str in ('X', 'None', ''):
                        current += timedelta(days=1)
                        continue
                    
                    daily.append({
                        'date': current,
                        'day': current.strftime('%a'),
                        'shift': shift_str,
                        'location': SHIFT_LEGEND.get(shift_str, shift_str),
                    })
                current += timedelta(days=1)
            
            if daily:
                staff_data.append({
                    'name': name_str,
                    'role': role,
                    'schedule': daily,
                })
        
        if staff_data:
            results[sheet_name] = {
                'start': str(effective_start),
                'end': str(effective_end),
                'staff': staff_data,
            }
    
    wb.close()
    
    if not results:
        return None, f"No staff schedule found for {start_date} to {end_date}"
    
    return results, None


def format_staff_for_email(data, period):
    """Build HTML for staff schedule email."""
    greeting = get_greeting()
    now = datetime.now()
    
    if period == "today":
        period_label = "Today"
    elif period == "week":
        period_label = "This Week"
    else:
        period_label = period
    
    # Determine date range
    all_dates = set()
    for team, tdata in data.items():
        for s in tdata['staff']:
            for d in s['schedule']:
                all_dates.add(d['date'])
    all_dates = sorted(all_dates)
    
    # Group by location and by off/vacation/leave
    OFF_CODES = {'V', 'LOA'}
    
    by_location = defaultdict(lambda: defaultdict(list))
    off_staff = defaultdict(lambda: defaultdict(list))
    for team, tdata in data.items():
        for s in tdata['staff']:
            for entry in s['schedule']:
                location = entry.get('location', entry['shift'])
                shift_code = entry['shift']
                
                if shift_code in OFF_CODES:
                    off_staff[shift_code][entry['date']].append({
                        'name': s['name'],
                        'role': s['role'],
                        'team': team,
                    })
                else:
                    by_location[location][entry['date']].append({
                        'name': s['name'],
                        'role': s['role'],
                        'team': team,
                    })
    
    date_sections = ""
    all_off_staff = defaultdict(lambda: defaultdict(list))  # code -> name -> [dates]
    for d in all_dates:
        date_str = d.strftime("%A, %B %d, %Y")
        
        # Build working staff sections grouped by location
        day_sections = ""
        for loc in sorted(by_location.keys()):
            staff_at_loc = by_location[loc].get(d, [])
            if not staff_at_loc:
                continue
            
            # Sort alphabetically by last name
            def sort_key(s):
                parts = s['name'].lower().strip().split(',', 1)
                return (parts[0].strip(), parts[1].strip()) if len(parts) == 2 else (s['name'].lower(), '')
            staff_at_loc.sort(key=sort_key)
            
            loc_name = loc
            staff_rows = ""
            for s in staff_at_loc:
                staff_rows += f"""
            <tr>
              <td style="padding:3px 8px;border-bottom:1px solid #e8e8e8;font-family:Georgia,'Times New Roman',serif;font-size:9pt;color:#333">{s['name']}</td>
              <td style="padding:3px 8px;border-bottom:1px solid #e8e8e8;font-family:Georgia,'Times New Roman',serif;font-size:9pt;color:#888">{s['role'] if s['role'] else '—'}</td>
            </tr>"""
            
            day_sections += f"""
      <tr><td style="padding:6px 0 2px 0;font-family:Georgia,'Times New Roman',serif;font-size:10pt;font-weight:bold;color:#1a3a5c">{loc_name}</td></tr>
      <tr><td style="padding:0 0 6px 0">
        <table cellpadding="0" cellspacing="0" width="100%" style="border-collapse:collapse">
          <tr style="color:#888;font-size:8pt;font-family:Georgia,'Times New Roman',serif">
            <td style="padding:3px 8px;border-bottom:2px solid #d0d4d8;font-weight:bold">Name</td>
            <td style="padding:3px 8px;border-bottom:2px solid #d0d4d8;width:60px;font-weight:bold">Role</td>
          </tr>
          {staff_rows}
        </table>
      </td></tr>"""
        
        # Collect off-staff for end-of-week summary (don't show per day)
        for code, label in [('V', 'Vacation'), ('LOA', 'Leave of Absence')]:
            off_list = off_staff[code].get(d, [])
            for s in off_list:
                all_off_staff[code][s['name']].append({
                    'date': d,
                    'role': s['role'],
                })
        
        if day_sections:
            date_sections += f"""
      <tr>
        <td style="padding:14px 0 4px 0;font-family:Georgia,'Times New Roman',serif;font-size:12pt;font-weight:bold;color:#1a3a5c">{date_str}</td>
      </tr>
      {day_sections}"""
    
    # Build off-staff summary section at the very end
    off_summary = ""
    for code, label in [('V', 'Vacation'), ('LOA', 'Leave of Absence')]:
        if code not in all_off_staff or not all_off_staff[code]:
            continue
        # Sort names
        sorted_names = sorted(all_off_staff[code].keys(), key=lambda n: n.split(',')[0].strip().lower() + ' ' + n.split(',')[1].strip().lower() if ',' in n else n.lower())
        off_rows = ""
        for name in sorted_names:
            date_entries = all_off_staff[code][name]
            # Get unique dates sorted
            unique_dates = sorted(set(e['date'] for e in date_entries))
            role = date_entries[0]['role']
            # Format date range or list
            if len(unique_dates) == 1:
                date_str = unique_dates[0].strftime("%b %d")
            else:
                # Check if consecutive
                consecutive = True
                for i in range(1, len(unique_dates)):
                    if (unique_dates[i] - unique_dates[i-1]).days != 1:
                        consecutive = False
                        break
                if consecutive:
                    date_str = f"{unique_dates[0].strftime('%b %d')} – {unique_dates[-1].strftime('%b %d')}"
                else:
                    date_str = ", ".join(d.strftime("%b %d") for d in unique_dates)
            
            off_rows += f"""
            <tr>
              <td style="padding:3px 8px;border-bottom:1px solid #ffd6d6;font-family:Georgia,'Times New Roman',serif;font-size:9pt;color:#333">{name}</td>
              <td style="padding:3px 8px;border-bottom:1px solid #ffd6d6;font-family:Georgia,'Times New Roman',serif;font-size:9pt;color:#888">{role if role else '—'}</td>
              <td style="padding:3px 8px;border-bottom:1px solid #ffd6d6;font-family:Georgia,'Times New Roman',serif;font-size:9pt;color:#c62828">{date_str}</td>
            </tr>"""
        
        if off_rows:
            off_summary += f"""
      <tr><td style="padding:14px 0 2px 0;font-family:Georgia,'Times New Roman',serif;font-size:11pt;font-weight:bold;color:#c62828">{label}</td></tr>
      <tr><td style="padding:0">
        <table cellpadding="0" cellspacing="0" width="100%" style="border-collapse:collapse">
          <tr style="color:#888;font-size:8pt;font-family:Georgia,'Times New Roman',serif">
            <td style="padding:3px 8px;border-bottom:2px solid #d0d4d8;font-weight:bold">Name</td>
            <td style="padding:3px 8px;border-bottom:2px solid #d0d4d8;width:60px;font-weight:bold">Role</td>
            <td style="padding:3px 8px;border-bottom:2px solid #d0d4d8;width:140px;font-weight:bold">Date(s)</td>
          </tr>
          {off_rows}
        </table>
      </td></tr>"""
    
    if off_summary:
        date_sections += f"""
      <tr>
        <td style="padding:18px 0 4px 0;font-family:Georgia,'Times New Roman',serif;font-size:13pt;font-weight:bold;color:#1a3a5c">Not Working This Week</td>
      </tr>
      {off_summary}"""
    
    html = f"""<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0"></head>
<body style="margin:0;padding:0;background-color:#f0f2f5;font-family:'Times New Roman',Georgia,serif">
<table cellpadding="0" cellspacing="0" width="100%" style="background-color:#f0f2f5">
  <tr><td style="padding:30px 10px" align="center">
    <table cellpadding="0" cellspacing="0" style="max-width:620px;width:100%;background-color:#ffffff;border-radius:4px;box-shadow:0 1px 3px rgba(0,0,0,0.08)">
      <tr>
        <td style="background-color:#1a3a5c;border-radius:4px 4px 0 0;padding:18px 28px">
          <table cellpadding="0" cellspacing="0" width="100%">
            <tr>
              <td style="font-family:Georgia,'Times New Roman',serif;font-size:16pt;font-weight:bold;color:#ffffff">Montefiore Einstein Urology</td>
              <td style="font-family:Georgia,'Times New Roman',serif;font-size:9pt;color:#b0c4de;text-align:right;vertical-align:bottom">{now.strftime('%B %d, %Y')}</td>
            </tr>
          </table>
        </td>
      </tr>
      <tr>
        <td style="padding:16px 28px 0 28px;font-family:Georgia,'Times New Roman',serif;font-size:16pt;font-weight:bold;color:#1a3a5c">Staff Schedule — {period_label}</td>
      </tr>
      <tr><td style="padding:0 28px"><hr style="border:none;border-top:1px solid #d0d4d8;margin:0"></td></tr>
      <tr>
        <td style="padding:14px 28px 0 28px">
          <p style="margin:0 0 4px 0;font-family:Times New Roman,Georgia,serif;font-size:12pt;color:#333;line-height:1.5">{greeting} Shareef,</p>
          <p style="margin:0 0 12px 0;font-family:Times New Roman,Georgia,serif;font-size:11pt;color:#333;line-height:1.5">Here is the staff schedule for <b>{period_label.lower()}</b>.</p>
        </td>
      </tr>
      {date_sections}
      <tr>
        <td style="padding:10px 28px 24px 28px">
          <hr style="border:none;border-top:1px solid #ddd;margin:0 0 10px 0">
          {SIG_BLOCK}
        </td>
      </tr>
    </table>
    <table cellpadding="0" cellspacing="0" style="max-width:620px;width:100%;padding-top:10px">
      <tr><td style="font-family:Arial,Helvetica,sans-serif;font-size:8pt;color:#aaa;text-align:center">Montefiore Urology &bull; 1250 Waters Place &bull; Bronx, NY 10461</td></tr>
    </table>
  </td></tr>
</table>
</body>
</html>"""
    return html


def generate_staff_pdf(data, period_label):
    """Generate PDF for staff schedule."""
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.lib.colors import HexColor
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    import tempfile
    
    pdf = tempfile.NamedTemporaryFile(suffix='.pdf', delete=False)
    pdf_path = pdf.name
    pdf.close()
    
    doc = SimpleDocTemplate(pdf_path, pagesize=letter,
        topMargin=0.5*inch, bottomMargin=0.5*inch,
        leftMargin=0.5*inch, rightMargin=0.5*inch)
    
    styles = getSampleStyleSheet()
    s_title = ParagraphStyle('Title', parent=styles['Title'],
        textColor=HexColor('#1a3a5c'), fontSize=16, spaceAfter=2, fontName='Helvetica-Bold')
    s_sub = ParagraphStyle('Sub', parent=styles['Normal'],
        textColor=HexColor('#666666'), fontSize=10, spaceAfter=12)
    s_body = ParagraphStyle('Body', parent=styles['Normal'],
        textColor=HexColor('#333333'), fontSize=8, leading=11, spaceAfter=2)
    s_muted = ParagraphStyle('Muted', parent=s_body, textColor=HexColor('#888888'), fontSize=7)
    s_hdr = ParagraphStyle('HdrCell', parent=s_body, textColor=HexColor('#555555'), fontSize=8)
    
    elements = []
    
    # Header
    hdr = Table([
        [Paragraph("<b>Montefiore Einstein Urology</b>", ParagraphStyle('Hdr', parent=s_title, fontSize=12, alignment=TA_LEFT)),
         Paragraph("Staff Schedule", ParagraphStyle('HdrDate', parent=s_sub, fontSize=10, textColor=HexColor('#ffffff'), alignment=TA_RIGHT))]
    ], colWidths=[4.5*inch, 3*inch])
    hdr.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), HexColor('#1a3a5c')),
        ('TEXTCOLOR', (0, 0), (-1, -1), HexColor('#ffffff')),
        ('LEFTPADDING', (0, 0), (0, 0), 10),
        ('RIGHTPADDING', (-1, -1), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(hdr)
    elements.append(Spacer(1, 10))
    
    now = datetime.now()
    elements.append(Paragraph(f"<b>Staff Schedule — {period_label}</b>", s_title))
    elements.append(Paragraph(f"Generated {now.strftime('%B %d, %Y')}", s_sub))
    
    # Group by date then team
    all_dates = set()
    for team, tdata in data.items():
        for s in tdata['staff']:
            for d in s['schedule']:
                all_dates.add(d['date'])
    all_dates = sorted(all_dates)
    
    # Group by location and by off/vacation/leave
    OFF_CODES = {'V', 'LOA'}
    
    by_location = defaultdict(lambda: defaultdict(list))
    off_staff = defaultdict(lambda: defaultdict(list))
    for team, tdata in data.items():
        for s in tdata['staff']:
            for entry in s['schedule']:
                location = entry.get('location', entry['shift'])
                shift_code = entry['shift']
                
                if shift_code in OFF_CODES:
                    off_staff[shift_code][entry['date']].append({
                        'name': s['name'],
                        'role': s['role'],
                        'team': team,
                    })
                else:
                    by_location[location][entry['date']].append({
                        'name': s['name'],
                        'role': s['role'],
                        'team': team,
                    })

    all_off_staff_pdf = defaultdict(lambda: defaultdict(list))

    for d in all_dates:
        elements.append(Paragraph(f"<b>{d.strftime('%A, %B %d, %Y')}</b>",
            ParagraphStyle('DayHdr', parent=s_body, fontSize=10, spaceBefore=10, spaceAfter=4, fontName='Helvetica-Bold', textColor=HexColor('#1a3a5c'))))
        
        for loc in sorted(by_location.keys()):
            staff_at_loc = by_location[loc].get(d, [])
            if not staff_at_loc:
                continue
            
            staff_at_loc.sort(key=lambda s: s['name'].lower().split(',')[::-1])
            
            txn_data = [[
                Paragraph("<b>Name</b>", s_muted),
                Paragraph("<b>Role</b>", s_muted),
            ]]
            for s in staff_at_loc:
                txn_data.append([
                    Paragraph(s['name'], s_body),
                    Paragraph(s['role'] if s['role'] else '—', s_muted),
                ])
            
            if len(txn_data) > 1:
                t = Table(txn_data, colWidths=[4.0*inch, 1.0*inch])
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), HexColor('#f4f6f9')),
                    ('GRID', (0, 0), (-1, -1), 0.25, HexColor('#e0e4e8')),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 4),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                    ('TOPPADDING', (0, 0), (-1, -1), 3),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                ]))
                elements.append(t)
                elements.append(Spacer(1, 4))
        
        # Collect off-staff for end-of-week summary (don't show per day)
        for code, label in [('V', 'Vacation'), ('LOA', 'Leave of Absence')]:
            off_list = off_staff[code].get(d, [])
            for s in off_list:
                all_off_staff_pdf[code][s['name']].append({
                    'date': d,
                    'role': s['role'],
                })
        
        if d != all_dates[-1]:
            elements.append(Spacer(1, 6))
    
    # Add off-staff summary at the end
    for code, label in [('V', 'Vacation'), ('LOA', 'Leave of Absence')]:
        if code not in all_off_staff_pdf or not all_off_staff_pdf[code]:
            continue
        sorted_names = sorted(all_off_staff_pdf[code].keys(), key=lambda n: n.split(',')[0].strip().lower() if ',' in n else n.lower())
        off_data = [[
            Paragraph("<b>Name</b>", s_muted),
            Paragraph("<b>Role</b>", s_muted),
            Paragraph(f"<b>{label} Dates</b>", ParagraphStyle('RedHdr', parent=s_muted, textColor=HexColor('#c62828'))),
        ]]
        for name in sorted_names:
            date_entries = all_off_staff_pdf[code][name]
            unique_dates = sorted(set(e['date'] for e in date_entries))
            role = date_entries[0]['role']
            if len(unique_dates) == 1:
                date_str = unique_dates[0].strftime("%b %d")
            else:
                consecutive = all((unique_dates[i] - unique_dates[i-1]).days == 1 for i in range(1, len(unique_dates)))
                if consecutive:
                    date_str = f"{unique_dates[0].strftime('%b %d')} – {unique_dates[-1].strftime('%b %d')}"
                else:
                    date_str = ", ".join(d.strftime("%b %d") for d in unique_dates)
            off_data.append([
                Paragraph(name, s_body),
                Paragraph(role if role else '—', s_muted),
                Paragraph(f"<font color='#c62828'>{date_str}</font>", s_body),
            ])
        if len(off_data) > 1:
            t = Table(off_data, colWidths=[2.5*inch, 0.8*inch, 1.7*inch])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), HexColor('#fff0f0')),
                ('GRID', (0, 0), (-1, -1), 0.25, HexColor('#e0e4e8')),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ]))
            elements.append(Spacer(1, 8))
            elements.append(Paragraph(f"<b>Not Working This Week — {label}</b>",
                ParagraphStyle('OffHdr', parent=s_body, fontSize=9, fontName='Helvetica-Bold', textColor=HexColor('#c62828'))))
            elements.append(t)
    
    elements.append(Spacer(1, 15))
    elements.append(Paragraph(
        "<i>Generated by Hermes Agent — Montefiore Urology</i>",
        ParagraphStyle('Footer', parent=s_body, textColor=HexColor('#999999'), fontSize=6.5, alignment=TA_CENTER)))
    
    doc.build(elements)
    return pdf_path


def send_staff_schedule(period, recipient):
    """Send staff schedule email with PDF attachment."""
    from google_workspace import GoogleWorkspace
    
    data, error = get_staff_schedule(period)
    if error:
        return f"❌ {error}"
    
    html = format_staff_for_email(data, period)
    
    now = datetime.now()
    if period == "today":
        period_label = f"Today ({now.strftime('%b %d')})"
    elif period == "week":
        period_label = f"Week of {now.strftime('%b %d')}"
    else:
        period_label = f"{now.strftime('%B %Y')}"
    
    pdf_path = generate_staff_pdf(data, period_label)
    
    ws = GoogleWorkspace()
    result = ws.send_email(
        user_id="urologyresidency",
        to=recipient,
        subject=f"Montefiore Urology — Staff Schedule ({period_label})",
        body=html,
        attachments=[pdf_path],
        from_email="sfrasier@montefiore.org",
        is_html=True,
    )
    
    if result.get("successful"):
        return f"✅ Staff schedule sent to {recipient}"
    else:
        return f"❌ Failed: {json.dumps(result, indent=2, default=str)[:300]}"


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--call", choices=["today", "week", "month"], help="Send call schedule")
    parser.add_argument("--staff", choices=["today", "week", "month"], help="Send staff schedule")
    parser.add_argument("--email", default=DEFAULT_RECIPIENT)
    args = parser.parse_args()
    
    if args.staff:
        print(f"\n📧 Sending staff schedule ({args.staff}) to {args.email}...")
        result = send_staff_schedule(args.staff, args.email)
        print(f"  {result}")
    elif args.call:
        print(f"\n📧 Sending call schedule ({args.call}) to {args.email}...")
        result = send_call_schedule(args.call, args.email)
        print(f"  {result}")
    else:
        print("Specify --call today|week|month or --staff today|week|month")
