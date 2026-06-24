#!/usr/bin/env python3
"""
Read reimbursement data from Resident_Trackers2025-2026.xlsx.
Single source of truth for all reimbursement reports and emails.

Data source: /workspace/Resident_Trackers2025-2026.xlsx
  - Summary sheet: per-resident GME totals, account balances
  - Sheet 2025-2026: individual transactions with cost center allocations

Account names displayed:
  - "Donation Account" → "MISC" (never shown as Donation)
"""

import openpyxl
from pathlib import Path
from collections import defaultdict, OrderedDict
import re

XLSX_PATH = Path("/workspace/Resident_Trackers2025-2026.xlsx")


def load_workbook():
    return openpyxl.load_workbook(str(XLSX_PATH), data_only=True)


def map_account(cost_center, comments=""):
    """Map cost center to display account name. Donation → MISC."""
    if not cost_center:
        cost_center = ""
    if not comments:
        comments = ""

    uc = cost_center.upper()
    if 'GME' in uc:
        return 'GME Funds'
    if 'TEACHING' in uc or cost_center == '130095':
        return 'Teaching Funds'
    if 'DEPT' in uc or 'INSTACART' in uc:
        return 'Dept Funds'
    if 'DONATION' in uc or cost_center == '100305095019':
        return 'MISC'
    if 'SLEEP' in uc:
        return 'Sleep Deprivation'
    if 'DONATION' in comments.upper() or 'DONATION' in str(comments).upper():
        return 'MISC'
    return 'Other'


def get_resident_summary():
    """Get per-resident GME totals from Summary sheet."""
    wb = load_workbook()
    ws = wb["Summary"]
    residents = {}
    for r in range(16, 45):
        name = ws.cell(r, 1).value
        charged = ws.cell(r, 2).value
        remaining = ws.cell(r, 3).value
        status = ws.cell(r, 4).value
        if name and charged is not None:
            name = str(name).strip()
            if name == 'Resident' or isinstance(charged, str):
                continue
            c = float(charged)
            residents[name.lower()] = {
                'name': name,
                'gme_used': c,
                'gme_remaining': float(remaining or 0),
                'gme_status': str(status or '').strip(),
                'annual_cap': 1250.0,
                'gme_pct': round(min(100, (c / 1250.0) * 100), 1),
            }
    wb.close()
    return residents


def parse_gme_amount(cost_center, full_amount, comments=""):
    """Parse the GME portion from a cost center string."""
    if not cost_center or 'GME' not in cost_center.upper():
        return 0.0

    if comments and 'GME' in comments.upper():
        has_remaining = 'remaining' in comments.lower()
        has_split_slash = '/' in comments and not has_remaining
        if has_split_slash or (not has_remaining and 'GME' in comments.upper()):
            gme_in_comments = re.search(r'GME[^$]*\$?\s*([\d,]+\.?\d*)', comments, re.IGNORECASE)
            if gme_in_comments:
                return float(gme_in_comments.group(1).replace(',', ''))

    if '/' not in cost_center:
        return float(full_amount)

    parts = cost_center.split('/')
    for part in parts:
        p = part.strip()
        if 'GME' in p.upper():
            dollar = re.search(r'\$?\s*([\d,]+\.?\d*)\s*GME', p, re.IGNORECASE)
            if dollar:
                return float(dollar.group(1).replace(',', ''))
            dollar2 = re.search(r'GME[^$]*\$?\s*([\d,]+\.?\d*)', p, re.IGNORECASE)
            if dollar2:
                amt = float(dollar2.group(1).replace(',', ''))
                if amt > 2000 and amt > 10000:
                    return amt / 10
                return amt
            other_parts = [x.strip() for x in parts if x.strip() != p]
            other_has_dollar = any('$' in x for x in other_parts)
            if other_has_dollar:
                other_total = 0
                for op in other_parts:
                    od = re.search(r'\$?\s*([\d,]+\.?\d*)', op)
                    if od:
                        other_total += float(od.group(1).replace(',', ''))
                remaining = float(full_amount) - other_total
                if remaining > 0 and remaining <= float(full_amount):
                    return remaining
            return float(full_amount) / len(parts)
    return float(full_amount)


def get_all_transactions(resident_name=None):
    """
    Get ALL transactions from Sheet 2025-2026, with account mapping.
    Returns dict of resident_name_lower -> list of {date, description, amount, account}
    When resident_name is provided, all name variants (typos) are merged into one key.
    """
    wb = load_workbook()
    ws = wb["Sheet 2025-2026"]
    by_resident = defaultdict(list)

    # Normalize resident_name for matching
    if resident_name:
        rn_base = resident_name.lower().replace(' ', '').replace(',', '')

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[0] or '').strip()
        if not name:
            continue
        nlower = name.lower()

        # Flexible name matching
        if resident_name:
            nn = nlower.replace(' ', '').replace(',', '')
            # Direct match
            if rn_base == nn or rn_base in nn or nn in rn_base:
                pass  # matched
            # Double-letter normalization (Kelli vs Keli)
            elif rn_base.replace('ll', 'l') == nn.replace('ll', 'l'):
                pass  # matched
            elif rn_base.replace('bb', 'b') == nn.replace('bb', 'b'):
                pass  # matched
            else:
                continue  # no match

        date_val = row[1]
        desc = str(row[3] or '').strip()
        amt = float(row[4] or 0)
        cc = str(row[7] or '').strip()
        comments = str(row[8] or '').strip()
        date_str = str(date_val)[:10] if date_val else ''

        if amt <= 0:
            continue

        # Filter by current academic year (July 1 → June 30)
        if date_val:
            try:
                from datetime import datetime
                dt = date_val if isinstance(date_val, datetime) else datetime.strptime(date_str[:10], '%Y-%m-%d')
                month, year = dt.month, dt.year
                
                # Handle obvious date typos (e.g. 2526 → 2026)
                if year > 2030 and dt.year < 3000:
                    year = 2000 + (year % 100)  # 2526 → 2026
                    # Also fix the display date_str
                    date_str = f"{year}-{month:02d}-{dt.day:02d}"
                
                # Current AY: Jul 2025 → Jun 2026
                if not (year == 2025 and month >= 7 or year == 2026 and month <= 6):
                    continue
            except (ValueError, TypeError):
                continue
        else:
            # Skip transactions with no date — can't determine AY
            continue

        account = map_account(cc, comments)
        
        # For GME and Teaching accounts, check if there's a split
        # The full amount may span multiple accounts
        actual_gme = parse_gme_amount(cc, amt, comments)
        
        if account == 'GME Funds':
            # Only show the actual GME portion under GME
            actual_amount = actual_gme
        elif 'GME' in cc.upper() and '/' in cc:
            # This is a split — GME portion goes to GME, remainder goes to source account
            # But the non-GME portion was already counted in the full amount
            # We need to NOT double count: show the item once with the actual account
            pass  # handled below
        else:
            actual_amount = amt
        
        # For split items (cost center has / and GME), the GME portion and Teaching/Dept portion
        # are separate line items
        if '/' in cc and 'GME' in cc.upper():
            gme_amt = actual_gme
            non_gme_amt = amt - gme_amt
            
            # Determine the non-GME account
            non_gme_account = 'Other'
            non_gme_cc = cc.split('/')[0].strip() if '/' in cc else ''
            if 'TEACHING' in non_gme_cc.upper() or 'TEACHING' in cc.upper():
                non_gme_account = 'Teaching Funds'
            elif 'DEPT' in non_gme_cc.upper():
                non_gme_account = 'Dept Funds'
            
            # Add GME portion
            if gme_amt > 0:
                by_resident[nlower].append({
                    'name': name,
                    'date': date_str,
                    'description': desc,
                    'amount': gme_amt,
                    'account': 'GME Funds',
                })
            
            # Add non-GME portion as separate line
            if non_gme_amt > 0:
                by_resident[nlower].append({
                    'name': name,
                    'date': date_str,
                    'description': desc,
                    'amount': non_gme_amt,
                    'account': non_gme_account,
                })
        else:
            by_resident[nlower].append({
                'name': name,
                'date': date_str,
                'description': desc,
                'amount': amt,
                'account': account,
            })

    wb.close()
    
    # When filtering by resident, merge all name variants into one entry
    if resident_name and len(by_resident) > 1:
        merged = {}
        all_txns = []
        for key, tlist in by_resident.items():
            all_txns.extend(tlist)
        # Use the first key found
        first_key = list(by_resident.keys())[0]
        merged[first_key] = all_txns
        return merged
    
    return dict(by_resident)


def get_account_totals(resident_name):
    """Get per-account totals for a resident."""
    txns = get_all_transactions(resident_name)
    totals = OrderedDict()
    total_all = 0
    for nlower, tlist in txns.items():
        by_acct = defaultdict(float)
        for t in tlist:
            by_acct[t['account']] += t['amount']
            total_all += t['amount']
        for acct in sorted(by_acct.keys()):
            totals[acct] = by_acct[acct]
    totals['Total'] = total_all
    return totals


if __name__ == "__main__":
    import sys
    target = sys.argv[1] if len(sys.argv) > 1 else "Kelli Aibel"
    
    # Show all transactions
    txns_by_res = get_all_transactions(target)
    summary = get_resident_summary()
    
    print(f"=== {target} — ALL TRANSACTIONS (AY 2025-2026) ===")
    for nlower, tlist in txns_by_res.items():
        for t in tlist:
            print(f"  {t['date']} | ${t['amount']:>7,.2f} | {t['account']:20s} | {t['description']}")
    
    # Account totals
    acct_totals = get_account_totals(target)
    print(f"\n=== BY ACCOUNT ===")
    for acct, total in acct_totals.items():
        print(f"  {acct:20s}: ${total:>8,.2f}")
    
    # GME summary
    for nlower in [target.lower(), target.lower().replace("kelli", "keli")]:
        if nlower in summary:
            r = summary[nlower]
            print(f"\n  GME: ${r['gme_used']:,.2f} used, ${r['gme_remaining']:,.2f} remaining — {r['gme_status']}")
            break
