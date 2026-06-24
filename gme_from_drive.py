#!/usr/bin/env python3
"""GME data from Drive spreadsheet. Outputs JSON."""
import openpyxl, re, json, sys
from pathlib import Path
from collections import defaultdict

CACHE = Path("/workspace/agentic-os/reports/gme_cache.json")
XLSX = Path("/workspace/reimbursement_tracker.xlsx")

def get_gme(cc, amt, comments):
    c = cc.upper().replace(" ", "")
    # Split entries like "Dept Funds $500 / GME Funds $1250"
    if "/" in c:
        for part in c.split("/"):
            p = part.strip()
            if "GME" in p or "10039" in p:
                # Search this specific part for a dollar amount
                m = re.search(r"\$?\s*([\d,]+\.?\d*)", p.replace("GMEFUNDS","").replace("(10039)",""))
                if m:
                    val = float(m.group(1).replace(",",""))
                    if val > 2000: val /= 10
                    return val
                # No $ in the GME part — search within this part specifically
                m = re.search(r"GME.*?(\d+\.?\d*)", p)
                if m:
                    return float(m.group(1))
                # Still nothing — check the comments for a GME amount
                m = re.search(r"GME\s*Funds?\s*\$?\s*([\d,]+\.?\d*)", comments, re.IGNORECASE)
                if m:
                    val = float(m.group(1).replace(",",""))
                    if val > 2000: val /= 10
                    return val
    # Direct GME-only cost center
    if "GME" in c and "DEPT" not in c and "TEACHING" not in c and "DONATION" not in c and "INSTACART" not in c:
        return float(amt)
    # Non-GME cost centers
    if c == "130095" or c.startswith("100305095019"): return 0.0
    if "TEACHING" in c and "GME" not in c: return 0.0
    if "DEPT" in c and "GME" not in c: return 0.0
    if "DONATION" in c and "GME" not in c: return 0.0
    return 0.0

def parse():
    wb = openpyxl.load_workbook(str(XLSX), read_only=True, data_only=True)
    if "Sheet 2025-2026" not in wb.sheetnames:
        wb.close()
        return {"error": "Sheet not found"}
    ws = wb["Sheet 2025-2026"]

    gme_by_name = defaultdict(float)
    for row in ws.iter_rows(values_only=True):
        cells = [str(c).strip() if c else "" for c in row]
        if not cells or not cells[0]: continue
        try: amt = float(cells[4].replace(",",""))
        except: continue
        if amt == 0: continue
        cc = cells[7] if len(cells) > 7 else ""
        comments = cells[8] if len(cells) > 8 else ""
        g = get_gme(cc, amt, comments)
        if g > 0:
            gme_by_name[cells[0].lower().strip()] += g
    wb.close()

    # Merge Kelli/Keli
    if "keli aibel" in gme_by_name:
        gme_by_name["kelli aibel"] = gme_by_name.get("kelli aibel", 0) + gme_by_name.pop("keli aibel")

    cap = 1250.0
    total_used = 0.0
    name_map = [
        ("ariel allen","Ariel Allen"), ("dimindra karki","Dimindra Karki"),
        ("dinora murota","Dinora Murota"), ("farzaan kassam","Farzaan Kassam"),
        ("jake drobner","Jake Drobner"), ("jasmin capellan","Jasmin Capellan"),
        ("john hill","John Hill"), ("john hordines","John Hordines"),
        ("joseph kim","Joseph Kim"), ("kelli aibel","Kelli Aibel"),
        ("nathaniel iskhakov","Nathaniel Iskhakov"), ("rutul patel","Rutul Patel"),
        ("samuel yim","Samuel Yim"), ("so yeon pak","So Yeon Pak"),
        ("valmic patel","Valmic Patel"),
    ]
    residents = []
    for key, display in name_map:
        gme = gme_by_name.get(key, 0)
        capped = min(gme, cap)
        rem = round(max(0, cap - capped), 2)
        pct = round((capped / cap) * 100, 1) if capped > 0 else 0
        total_used += capped
        residents.append({"name": display, "cls": "", "gme_used": round(capped, 2),
                          "annual_cap": cap, "remaining": rem, "usage_pct": pct})

    result = {
        "annual_cap_per_resident": cap,
        "total_residents": len(residents),
        "total_cap": round(len(residents) * cap, 2),
        "total_used": round(total_used, 2),
        "total_remaining": round(max(0, (len(residents) * cap) - total_used), 2),
        "residents": residents, "source": "drive_spreadsheet",
    }
    CACHE.parent.mkdir(parents=True, exist_ok=True)
    CACHE.write_text(json.dumps(result))
    return result

if __name__ == "__main__":
    print(json.dumps(parse(), indent=2))
